"""
extraer_proyectos.py
══════════════════════════════════════════════════════════════════════
Scraper de proyectos de ley de la Asamblea Legislativa de Costa Rica.
 
Qué hace
─────────
1. Abre el portal SIL con Playwright (Chromium).
2. Navega por todas las páginas de resultados (o un límite definido).
3. Por cada proyecto: abre el panel de detalle y extrae:
     • Datos maestros   (tipo, fechas, gaceta, número de ley)
     • Proponentes      (secuencia, apellidos, nombre)
     • Tramitación      (órgano, fechas, tipo de trámite)
     • Documento        (PDF firmado o formato editable, opcional)
4. Llama a sync_engine para persistir los datos en PostgreSQL.
5. Guarda un backup en JSON + Excel con tres hojas.
 
Configuración rápida
─────────────────────
  MAX_PAGINAS    = None  → recorre TODAS las páginas
  MAX_PAGINAS    = 5     → solo las primeras 5 páginas (≈50 proyectos)
  DESCARGAR_DOCS = True  → descarga PDFs/DOCX a la carpeta CARPETA_DOCS
  DESCARGAR_DOCS = False → omite descargas (útil en CI/CD para ahorrar tiempo)
 
Dependencias
─────────────
  pip install playwright openpyxl
  playwright install chromium
 
Variables de entorno
─────────────────────
  DATABASE_URL → ver sync_engine.py
"""
 
import asyncio
import json
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from playwright.async_api import async_playwright
 
from sync_engine import crear_tablas, sync_proyectos
 
# ══════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════════════
 
URL = (
    "https://www.asamblea.go.cr/Centro_de_informacion/"
    "Consultas_SIL/SitePages/ConsultaProyectos.aspx"
)
 
CARPETA_DOCS   = "documentos"   # carpeta local donde se guardan los archivos
MAX_PAGINAS = int(os.getenv("MAX_PAGINAS", "1")) or None           # None = todas las páginas | int = límite
DESCARGAR_DOCS = False          # True activa descargas de PDF/DOCX
 
 
# ══════════════════════════════════════════════════════════════════════
# UTILIDADES DE TEXTO
# ══════════════════════════════════════════════════════════════════════
 
def limpiar_texto(valor):
    """
    Elimina caracteres que openpyxl no acepta en celdas de Excel.
 
    Problemas comunes del scraper:
      • Caracteres de control ASCII (0x00–0x1F): tabulaciones internas,
        saltos de línea embebidos, null bytes, etc.
      • Bloque "Control Pictures" (U+2400–U+243F): ␦ y similares que
        aparecen cuando el sitio devuelve placeholders de carácter vacío.
 
    Args:
        valor: Cualquier valor. Si no es str, se devuelve sin cambios.
 
    Returns:
        Cadena limpia apta para Excel, o el valor original si no es str.
    """
    if not isinstance(valor, str):
        return valor
    # Elimina control ASCII salvo \t y \n que openpyxl sí tolera
    valor = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', valor)
    # Reemplaza símbolos del bloque Control Pictures por un guion legible
    valor = re.sub(r'[\u2400-\u243F]', '-', valor)
    return valor
 
 
# ══════════════════════════════════════════════════════════════════════
# HELPERS DE FRAME / BOTONES
# ══════════════════════════════════════════════════════════════════════
 
async def encontrar_frame_con_proyectos(page):
    """
    Busca y devuelve el frame (o la página principal) que contiene
    la tabla de proyectos.
 
    El portal SIL puede embeber el contenido en un <iframe>. Esta función
    prueba primero la página raíz y luego cada frame hijo hasta encontrar
    uno que tenga botones de "Seleccionar".
 
    Returns:
        Frame/Page con proyectos, o None si no se encuentra nada.
    """
    selectores_boton = [
        "img[src*='seleccionar']",
        "img[alt*='eleccionar']",
        "input[value='Seleccionar']",
    ]
 
    # Intentar en la página raíz primero (caso más común)
    for sel in selectores_boton:
        if await page.query_selector_all(sel):
            return page
 
    # Buscar en frames hijo
    for frame in page.frames:
        try:
            for sel in selectores_boton + ["input[type='image']"]:
                if await frame.query_selector_all(sel):
                    return frame
        except Exception:
            continue  # frame destruido o sin acceso
 
    return None
 
 
async def get_botones_visibles(frame):
    """
    Devuelve la lista de botones de "Seleccionar" visibles en el frame.
 
    Se prueban varios selectores porque el portal no es consistente entre
    versiones: a veces usa <img>, a veces <input type="image">.
 
    Returns:
        Lista (posiblemente vacía) de ElementHandles visibles.
    """
    for sel in [
        "img[src*='seleccionar']",
        "img[alt*='eleccionar']",
        "input[type='image']",
        "input[value='Seleccionar']",
    ]:
        candidatos = await frame.query_selector_all(sel)
        visibles   = [el for el in candidatos if await el.is_visible()]
        if visibles:
            return visibles
    return []
 
 
def limpia(vals: list[str]) -> bool:
    """
    Retorna True si ningún valor contiene tabulaciones ni saltos de línea.
    Se usa para descartar filas de encabezado o celdas combinadas que
    Playwright extrae con caracteres de control embebidos.
    """
    return all("\t" not in v and "\n" not in v for v in vals)
 
 
# ══════════════════════════════════════════════════════════════════════
# EXTRACCIÓN DE DATOS DEL PANEL DE DETALLE
# ══════════════════════════════════════════════════════════════════════
 
async def extraer_detalle(frame) -> dict:
    """
    Extrae los campos maestros del proyecto desde el panel de detalle.
 
    Busca la tabla que contenga simultáneamente "Tipo de Expediente" y
    "Vencimiento" (identificadores únicos de esa sección). Luego lee
    pares clave-valor de las celdas.
 
    Returns:
        Dict con campos como:
            {
              "Tipo de Expediente": "...",
              "Fecha de Inicio": "25-mar.-2026",
              "Vencimiento Cuatrienal": "...",
              "Fecha de Publicación": "...",
              "Número de Gaceta": "...",
              "Número de Ley": "..."
            }
        Dict vacío si no se encuentra la tabla.
    """
    detalle = {}
    try:
        for tabla in await frame.query_selector_all("table"):
            txt = await tabla.inner_text()
            # Identificar la tabla correcta por su contenido
            if "Tipo de Expediente" in txt and "Vencimiento" in txt:
                for fila in await tabla.query_selector_all("tr"):
                    celdas = await fila.query_selector_all("td")
                    if len(celdas) == 2:
                        clave  = (await celdas[0].inner_text()).strip()
                        valor  = (await celdas[1].inner_text()).strip()
                        # Filtrar filas con encabezados multilínea o vacíos
                        if clave and "\t" not in clave and "\n" not in clave and len(clave) < 60:
                            detalle[clave] = valor
                break  # tabla encontrada, no seguir buscando
    except Exception as exc:
        print(f"    ⚠️  extraer_detalle: {exc}")
    return detalle
 
 
async def extraer_proponentes(frame) -> list:
    """
    Extrae la tabla de proponentes/firmantes del proyecto.
 
    Identifica la tabla correcta buscando los encabezados
    "Secuencia de Firma", "Apellidos" y "Nombre" en la primera fila.
 
    Returns:
        Lista de dicts:
            [ {"Secuencia": "1", "Apellidos": "PÉREZ MORA", "Nombre": "Juan"} ]
        Lista vacía si no se encuentra la tabla o está sin datos.
    """
    proponentes = []
    try:
        for tabla in await frame.query_selector_all("table"):
            filas = await tabla.query_selector_all("tr")
            if not filas:
                continue
            # Leer encabezados de la primera fila
            encabezados = [
                (await c.inner_text()).strip()
                for c in await filas[0].query_selector_all("td, th")
            ]
            if "Secuencia de Firma" in encabezados and "Apellidos" in encabezados and "Nombre" in encabezados:
                for fila in filas[1:]:
                    celdas = await fila.query_selector_all("td")
                    if len(celdas) >= 3:
                        vals = [(await c.inner_text()).strip() for c in celdas[:3]]
                        if limpia(vals) and any(vals):
                            proponentes.append({
                                "Secuencia": vals[0],
                                "Apellidos": vals[1],
                                "Nombre":    vals[2],
                            })
                break
    except Exception as exc:
        print(f"    ⚠️  extraer_proponentes: {exc}")
    return proponentes
 
 
async def extraer_tramitacion(frame) -> list:
    """
    Extrae el historial de tramitación del proyecto.
 
    Identifica la tabla buscando columnas que contengan "rgano" (Órgano)
    y "mite" (Trámite) como substrings, para tolerar variaciones de
    acentuación del portal.
 
    Returns:
        Lista de dicts:
            [
              {
                "Órgano":         "Comisión de Asuntos Jurídicos",
                "Fecha Inicio":   "01-feb.-2026",
                "Fecha Término":  "15-mar.-2026",
                "Tipo de Trámite":"Aprobación"
              }
            ]
        Lista vacía si no se encuentra la tabla.
    """
    tramitacion = []
    try:
        for tabla in await frame.query_selector_all("table"):
            filas = await tabla.query_selector_all("tr")
            if not filas:
                continue
            encabezados = [
                (await c.inner_text()).strip()
                for c in await filas[0].query_selector_all("td, th")
            ]
            # Búsqueda flexible de columnas (tolera acentos y variaciones)
            if any("rgano" in t for t in encabezados) and any("mite" in t for t in encabezados):
                for fila in filas[1:]:
                    celdas = await fila.query_selector_all("td")
                    if len(celdas) >= 4:
                        vals = [(await c.inner_text()).strip() for c in celdas[:4]]
                        if limpia(vals) and vals[0]:  # vals[0] = Órgano, no puede estar vacío
                            tramitacion.append({
                                "Órgano":         vals[0],
                                "Fecha Inicio":   vals[1],
                                "Fecha Término":  vals[2],
                                "Tipo de Trámite":vals[3],
                            })
                break
    except Exception as exc:
        print(f"    ⚠️  extraer_tramitacion: {exc}")
    return tramitacion
 
 
async def descargar_documento(page, frame, num_exp: str) -> dict:
    """
    Descarga el documento adjunto al proyecto (PDF firmado o DOCX editable).
 
    Prioridad de descarga:
        1. PDF Firmado  → extensión .pdf
        2. Formato Editable (Word) → extensión .docx
 
    Los botones de descarga se identifican buscando el texto de su fila
    padre. Si no se puede identificar, se toma el último botón visible
    como PDF y el primero como Word.
 
    Args:
        page:    Página Playwright (para capturar el evento de descarga).
        frame:   Frame activo donde están los botones.
        num_exp: Número de expediente, usado para nombrar el archivo.
 
    Returns:
        Dict con:
            { "archivo": "documentos/documento_12345.pdf", "tipo": "pdf" }
        o
            { "archivo": None, "tipo": "ninguno" }  si DESCARGAR_DOCS=False
        o si no hay documento disponible.
    """
    resultado = {"archivo": None, "tipo": "ninguno"}
 
    if not DESCARGAR_DOCS:
        print("     ⛔ Descarga deshabilitada (DESCARGAR_DOCS=False)")
        return resultado
 
    os.makedirs(CARPETA_DOCS, exist_ok=True)
 
    try:
        botones = await frame.query_selector_all(
            "input[value='Descargar'], button:has-text('Descargar')"
        )
        boton_pdf = boton_word = None
 
        # Identificar cada botón por el texto de su fila padre (<tr>)
        for boton in botones:
            try:
                texto_fila = await boton.evaluate("""el => {
                    let tr = el.closest('tr');
                    return tr ? tr.innerText : (el.parentElement ? el.parentElement.innerText : '');
                }""")
                texto_fila = texto_fila.strip().lower()
                if "pdf firmado" in texto_fila:
                    boton_pdf = boton
                elif "formato editable" in texto_fila or "word" in texto_fila:
                    boton_word = boton
            except Exception:
                pass
 
        # Fallback: si no se pudo identificar, asumir orden convencional
        if boton_pdf is None and boton_word is None:
            if len(botones) >= 2:
                boton_word, boton_pdf = botones[0], botones[1]
            elif len(botones) == 1:
                boton_word = botones[0]
 
        # Seleccionar el mejor botón disponible (PDF tiene prioridad)
        boton_usar = tipo = ext = None
        if boton_pdf and await boton_pdf.is_visible():
            boton_usar, tipo, ext = boton_pdf, "pdf", ".pdf"
            print("     📄 Descargando PDF Firmado...")
        elif boton_word and await boton_word.is_visible():
            boton_usar, tipo, ext = boton_word, "word", ".docx"
            print("     📝 Descargando Formato Editable (DOCX)...")
 
        if boton_usar:
            ruta = os.path.join(CARPETA_DOCS, f"documento_{num_exp}{ext}")
            async with page.expect_download(timeout=30_000) as dl:
                await boton_usar.click()
            await (await dl.value).save_as(ruta)
            resultado = {"archivo": ruta, "tipo": tipo}
            print(f"     ✅ Archivo guardado: {ruta}")
        else:
            print("     ℹ️  Sin documentos disponibles para descargar.")
 
    except Exception as exc:
        print(f"     ⚠️  Error al descargar documento del exp. {num_exp}: {exc}")
 
    return resultado
 
 
# ══════════════════════════════════════════════════════════════════════
# NAVEGACIÓN DE PÁGINAS
# ══════════════════════════════════════════════════════════════════════
 
async def obtener_links_paginacion(frame) -> dict:
    """
    Lee la barra de paginación y devuelve un dict de links disponibles.
 
    Escanea todos los <a> del frame y clasifica los que tienen texto
    puramente numérico o que son controles especiales de paginación.
 
    Returns:
        Dict donde las claves son:
          • int → número de página (ej: 1, 2, 11, 12)
          • str → control especial ("<<", ">>", "...", "Anterior", "Siguiente")
        y los valores son los ElementHandles correspondientes.
 
    Ejemplo de resultado:
        { 1: <el>, 2: <el>, ..., 10: <el>, "...": <el>, ">>": <el> }
    """
    resultado = {}
    controles_especiales = {"<<", ">>", "<", ">", "...", "Anterior", "Siguiente"}
    try:
        for link in await frame.query_selector_all("a"):
            try:
                texto = (await link.inner_text()).strip()
            except Exception:
                continue
            if texto.isdigit():
                resultado[int(texto)] = link
            elif texto in controles_especiales:
                resultado[texto] = link
    except Exception:
        pass
    return resultado
 
 
async def ir_a_pagina(frame, page, numero_pagina: int) -> bool:
    """
    Navega a la página indicada usando la barra de paginación del portal.
 
    El portal SIL muestra bloques de 10 páginas: [1..10], luego [11..20], etc.
    Para llegar a la página 11 desde la 10 hay que hacer clic en "..." o en
    el número 11 si ya es visible.
 
    Estrategia (en orden de prioridad):
        1. Si el número exacto está visible → clic directo.
        2. Si hay "..." → clic para avanzar al siguiente bloque.
        3. Si hay ">>" → clic como fallback.
        4. Sin opciones → devuelve False (fin de paginación).
 
    Se permiten hasta 5 intentos antes de rendirse.
 
    Args:
        frame:         Frame con la barra de paginación.
        page:          Página Playwright para esperar estados de carga.
        numero_pagina: Número de página al que se quiere ir.
 
    Returns:
        True  si se logró (o se asume que se logró) navegar a la página.
        False si no hay forma de avanzar (fin del sitio).
    """
    MAX_INTENTOS = 5
 
    for intento in range(1, MAX_INTENTOS + 1):
        links = await obtener_links_paginacion(frame)
        nums_visibles = sorted(k for k in links if isinstance(k, int))
        print(f"   📄 Paginación visible: {nums_visibles} (intento {intento}/{MAX_INTENTOS})")
 
        # ── Caso 1: número exacto visible → clic directo ──────────────
        if numero_pagina in links:
            try:
                await links[numero_pagina].click()
                await page.wait_for_load_state("domcontentloaded")
                await page.wait_for_timeout(2_000)
                print(f"   ✓ Navegado a página {numero_pagina} directamente.")
                return True
            except Exception as exc:
                print(f"   ⚠️  Error al clicar página {numero_pagina}: {exc}")
                return False
 
        # ── Caso 2: "..." → avanza al siguiente bloque ─────────────────
        if "..." in links:
            print(f"   → Página {numero_pagina} no visible, usando '...' para cambiar bloque.")
            try:
                await links["..."].click()
                await page.wait_for_load_state("domcontentloaded")
                await page.wait_for_timeout(2_000)
                print(f"   ✓ Bloque avanzado con '...'. Se asume llegada a página {numero_pagina}.")
                return True
            except Exception as exc:
                print(f"   ⚠️  Error al clicar '...': {exc}")
                return False
 
        # ── Caso 3: ">>" → fallback ────────────────────────────────────
        if ">>" in links:
            print(f"   → Usando '>>' como fallback para llegar a página {numero_pagina}.")
            try:
                await links[">>"].click()
                await page.wait_for_load_state("domcontentloaded")
                await page.wait_for_timeout(2_000)
                # Volver a buscar el frame actualizado en el siguiente intento
                frame = await encontrar_frame_con_proyectos(page)
                continue
            except Exception as exc:
                print(f"   ⚠️  Error al clicar '>>': {exc}")
                return False
 
        # ── Caso 4: sin opciones → fin de paginación ──────────────────
        print(f"   ✗ No hay forma de llegar a página {numero_pagina}. Fin de paginación.")
        return False
 
    print(f"   ✗ Se agotaron los {MAX_INTENTOS} intentos para llegar a página {numero_pagina}.")
    return False
 
 
# ══════════════════════════════════════════════════════════════════════
# PROCESAMIENTO DE PÁGINA
# ══════════════════════════════════════════════════════════════════════
 
async def procesar_pagina(page, frame_activo, num_pagina: int, todos_los_proyectos: list):
    """
    Extrae todos los proyectos visibles en la página actual.
 
    Flujo por proyecto:
        1. Leer número de expediente y título de la fila de la tabla.
        2. Hacer clic en el botón "Seleccionar" para abrir el panel de detalle.
        3. Esperar 2s y extraer detalle, proponentes, tramitación y documento.
        4. Agregar el dict resultado a todos_los_proyectos.
 
    Los índices de botones se recalculan en cada iteración porque el DOM
    puede cambiar después de abrir y cerrar paneles de detalle.
 
    Args:
        page:                 Página Playwright activa.
        frame_activo:         Frame que contiene la tabla de proyectos.
        num_pagina:           Número de página (solo para logging).
        todos_los_proyectos:  Lista acumulada (se modifica in-place).
    """
    botones = await get_botones_visibles(frame_activo)
    if not botones:
        print(f"  ⚠️  No se encontraron botones en página {num_pagina}.")
        return
 
    total = len(botones)
    print(f"\n{'═'*60}")
    print(f"  PÁGINA {num_pagina} — {total} proyecto(s) encontrado(s)")
    print(f"{'═'*60}")
 
    for i in range(total):
        # Re-obtener botones en cada iteración (el DOM cambia entre clics)
        botones_actuales = await get_botones_visibles(frame_activo)
        if i >= len(botones_actuales):
            print(f"  ⚠️  Botón {i+1} ya no está disponible, saltando.")
            break
 
        boton     = botones_actuales[i]
        num_exp   = "?"
        titulo    = ""
 
        # Intentar leer expediente y título desde la fila de la tabla
        try:
            fila_handle = await boton.evaluate_handle("el => el.closest('tr')")
            celdas      = await fila_handle.as_element().query_selector_all("td")
            textos      = [(await c.inner_text()).strip() for c in celdas]
            textos      = [t for t in textos if t and "\t" not in t and "\n" not in t]
            # El expediente es la primera celda puramente numérica
            for t in textos:
                if t.isdigit():
                    num_exp = t
                    break
            # El título es la celda con más texto
            if textos:
                titulo = max(textos, key=len)
        except Exception:
            pass  # Si falla, se usarán los defaults "?" y ""
 
        titulo_display = limpiar_texto(titulo)
        titulo_corto   = titulo_display[:60] + ("..." if len(titulo_display) > 60 else "")
        print(f"\n  📋 [{i+1}/{total}] Exp. {num_exp} — {titulo_corto}")
 
        # Abrir panel de detalle
        try:
            await boton.click()
            await page.wait_for_timeout(2_000)
        except Exception as exc:
            print(f"    ⚠️  Error al abrir panel de detalle: {exc}")
            continue
 
        # Extraer todos los datos del panel
        detalle     = await extraer_detalle(frame_activo)
        proponentes = await extraer_proponentes(frame_activo)
        tramitacion = await extraer_tramitacion(frame_activo)
        doc_info    = await descargar_documento(page, frame_activo, num_exp)
 
        todos_los_proyectos.append({
            "pagina":            num_pagina,
            "numero_expediente": num_exp,
            "titulo":            titulo,        # título original sin limpiar (para JSON)
            "detalle":           detalle,
            "proponentes":       proponentes,
            "tramitacion":       tramitacion,
            "documento":         doc_info,
        })
 
        print(
            f"    ✓ {len(detalle)} campo(s) | "
            f"{len(proponentes)} proponente(s) | "
            f"{len(tramitacion)} trámite(s) | "
            f"doc: {doc_info['tipo']}"
        )
 
 
# ══════════════════════════════════════════════════════════════════════
# EXPORTACIÓN A EXCEL
# ══════════════════════════════════════════════════════════════════════
 
def exportar_excel(todos_los_proyectos: list, nombre_archivo: str):
    """
    Guarda los proyectos en un archivo Excel (.xlsx) con tres hojas:
        • Resumen       → una fila por proyecto
        • Tramitación   → una fila por trámite
        • Proponentes   → una fila por proponente
 
    Aplica limpiar_texto() a todos los valores antes de escribirlos
    para evitar errores de openpyxl con caracteres ilegales.
 
    Args:
        todos_los_proyectos: Lista de proyectos ya extraídos.
        nombre_archivo:      Ruta completa del .xlsx a crear.
    """
    # Helpers de estilo
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    centrado    = Alignment(horizontal="center")
 
    def escribir_encabezado(ws, columnas: list):
        """Pinta la primera fila con el estilo de encabezado."""
        for col, texto in enumerate(columnas, start=1):
            celda            = ws.cell(row=1, column=col, value=texto)
            celda.font       = header_font
            celda.fill       = header_fill
            celda.alignment  = centrado
 
    def s(val):
        """Alias corto de limpiar_texto para uso interno."""
        return limpiar_texto(val) if isinstance(val, str) else val
 
    wb = openpyxl.Workbook()
 
    # ── Hoja 1: Resumen ────────────────────────────────────────────────
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    escribir_encabezado(ws_resumen, [
        "Página", "Expediente", "Título", "Tipo",
        "Fecha Inicio", "Vencimiento Cuatrienal", "Proponentes", "Documento"
    ])
    for fila_num, proy in enumerate(todos_los_proyectos, start=2):
        det  = proy["detalle"]
        # Concatenar proponentes como "APELLIDOS Nombre, ..."
        props_str = ", ".join(
            f"{p['Apellidos']} {p['Nombre']}".strip()
            for p in proy["proponentes"]
        )
        doc      = proy["documento"]
        doc_str  = f"{doc['archivo']} ({doc['tipo']})" if doc["archivo"] else "sin documento"
 
        ws_resumen.cell(row=fila_num, column=1, value=proy["pagina"])
        ws_resumen.cell(row=fila_num, column=2, value=s(proy["numero_expediente"]))
        ws_resumen.cell(row=fila_num, column=3, value=s(proy["titulo"]))
        ws_resumen.cell(row=fila_num, column=4, value=s(det.get("Tipo de Expediente", "")))
        ws_resumen.cell(row=fila_num, column=5, value=s(det.get("Fecha de Inicio", "")))
        ws_resumen.cell(row=fila_num, column=6, value=s(det.get("Vencimiento Cuatrienal", "")))
        ws_resumen.cell(row=fila_num, column=7, value=s(props_str))
        ws_resumen.cell(row=fila_num, column=8, value=s(doc_str))
 
    ws_resumen.column_dimensions["B"].width = 14
    ws_resumen.column_dimensions["C"].width = 65
    ws_resumen.column_dimensions["D"].width = 38
    ws_resumen.column_dimensions["G"].width = 45
    ws_resumen.column_dimensions["H"].width = 35
 
    # ── Hoja 2: Tramitación ────────────────────────────────────────────
    ws_tram = wb.create_sheet("Tramitación")
    escribir_encabezado(ws_tram, [
        "Expediente", "Órgano", "Fecha Inicio", "Fecha Término", "Tipo de Trámite"
    ])
    fila = 2
    for proy in todos_los_proyectos:
        for t in proy["tramitacion"]:
            ws_tram.cell(row=fila, column=1, value=s(proy["numero_expediente"]))
            ws_tram.cell(row=fila, column=2, value=s(t["Órgano"]))
            ws_tram.cell(row=fila, column=3, value=s(t["Fecha Inicio"]))
            ws_tram.cell(row=fila, column=4, value=s(t["Fecha Término"]))
            ws_tram.cell(row=fila, column=5, value=s(t["Tipo de Trámite"]))
            fila += 1
    ws_tram.column_dimensions["B"].width = 20
    ws_tram.column_dimensions["E"].width = 55
 
    # ── Hoja 3: Proponentes ────────────────────────────────────────────
    ws_prop = wb.create_sheet("Proponentes")
    escribir_encabezado(ws_prop, ["Expediente", "Secuencia", "Apellidos", "Nombre"])
    fila = 2
    for proy in todos_los_proyectos:
        for prop in proy["proponentes"]:
            ws_prop.cell(row=fila, column=1, value=s(proy["numero_expediente"]))
            ws_prop.cell(row=fila, column=2, value=s(prop["Secuencia"]))
            ws_prop.cell(row=fila, column=3, value=s(prop["Apellidos"]))
            ws_prop.cell(row=fila, column=4, value=s(prop["Nombre"]))
            fila += 1
 
    wb.save(nombre_archivo)
    print(f"📊 Excel guardado: {nombre_archivo}")
 
 
# ══════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════
 
async def main():
    inicio = datetime.now()
 
    print("╔══════════════════════════════════════════════════════════╗")
    print("║   Extractor de Proyectos — Asamblea Legislativa CR       ║")
    print("╚══════════════════════════════════════════════════════════╝")
    print(f"  Inicio         : {inicio.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Modo páginas   : {'TODAS' if not MAX_PAGINAS else f'primeras {MAX_PAGINAS}'}")
    print(f"  Descargar docs : {'Sí' if DESCARGAR_DOCS else 'No'}")
    print(f"  Carpeta docs   : {CARPETA_DOCS}/")
    print("=" * 60)
 
    os.makedirs(CARPETA_DOCS, exist_ok=True)
    todos_los_proyectos = []
 
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-blink-features=AutomationControlled",
            ]
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 800},
        )
        page = await context.new_page()

        print("\n📡 Cargando portal SIL...")
        await page.goto(URL, wait_until="domcontentloaded", timeout=60_000)
        await page.wait_for_timeout(8_000)  # un poco más que 6s para CI
 
        frame_activo = await encontrar_frame_con_proyectos(page)
        if not frame_activo:
            # Guardar screenshot para ver qué está recibiendo el browser
            await page.screenshot(path="debug_sin_frame.png", full_page=True)
            print("❌ No se encontró el frame con proyectos. Screenshot guardado.")
            await browser.close()
            return
 
        print("✅ Frame con proyectos localizado.\n")
 
        pagina_actual      = 1
        paginas_procesadas = set()
 
        while True:
            # Verificar límite de páginas
            if MAX_PAGINAS and pagina_actual > MAX_PAGINAS:
                print(f"\n🛑 Límite de {MAX_PAGINAS} página(s) alcanzado.")
                break
 
            # Evitar procesar la misma página dos veces (loop detectado)
            if pagina_actual in paginas_procesadas:
                print(f"\n⚠️  Página {pagina_actual} ya fue procesada. Posible loop detectado, deteniendo.")
                break
 
            await procesar_pagina(page, frame_activo, pagina_actual, todos_los_proyectos)
            paginas_procesadas.add(pagina_actual)
 
            # Intentar navegar a la siguiente página
            siguiente = pagina_actual + 1
            print(f"\n🔄 Navegando a página {siguiente}...")
            pudo_navegar = await ir_a_pagina(frame_activo, page, siguiente)
 
            if not pudo_navegar:
                print("\n✅ No hay más páginas disponibles. Extracción completada.")
                break
 
            pagina_actual = siguiente
 
        await browser.close()
 
    # ── Sin resultados ─────────────────────────────────────────────────
    if not todos_los_proyectos:
        print("\n⚠️  No se extrajeron proyectos. Revisá la URL o el estado del sitio.")
        return
 
    # ── Sync a PostgreSQL ──────────────────────────────────────────────
    print("\n" + "═" * 60)
    print("  SYNC A BASE DE DATOS")
    print("═" * 60)
    crear_tablas()
    stats = sync_proyectos(todos_los_proyectos)
 
    # ── Guardar archivos de backup ─────────────────────────────────────
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_file  = f"proyectos_{timestamp}.json"
    excel_file = f"proyectos_{timestamp}.xlsx"
 
    print("\n" + "═" * 60)
    print("  EXPORTANDO ARCHIVOS")
    print("═" * 60)
 
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(todos_los_proyectos, f, ensure_ascii=False, indent=2)
    print(f"💾 JSON guardado  : {json_file}")
 
    exportar_excel(todos_los_proyectos, excel_file)
 
    # ── Resumen final ──────────────────────────────────────────────────
    duracion = datetime.now() - inicio
    docs_ok  = sum(1 for p in todos_los_proyectos if p["documento"]["archivo"])
 
    print("\n" + "╔" + "═" * 58 + "╗")
    print("║   RESUMEN FINAL                                          ║")
    print("╠" + "═" * 58 + "╣")
    print(f"║  Proyectos extraídos : {len(todos_los_proyectos):<33}║")
    print(f"║  Páginas procesadas  : {str(sorted(paginas_procesadas)):<33}║")
    print(f"║  Documentos descarg. : {docs_ok}/{len(todos_los_proyectos):<31}║")
    print(f"║  DB — nuevos         : {stats['nuevos']:<33}║")
    print(f"║  DB — duplicados     : {stats['duplicados']:<33}║")
    print(f"║  DB — errores        : {stats['errores']:<33}║")
    print(f"║  JSON                : {json_file:<33}║")
    print(f"║  Excel               : {excel_file:<33}║")
    print(f"║  Duración total      : {str(duracion).split('.')[0]:<33}║")
    print("╚" + "═" * 58 + "╝")
 
 
if __name__ == "__main__":
    asyncio.run(main())