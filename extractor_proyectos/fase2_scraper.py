"""
fase2_scraper.py
----------------------------------------------------------------------
Fase 2 - Backfill y actualización continua del SIL.

Responsabilidad: recorrer TODAS las páginas del portal de forma
gradual, guardando el progreso en PostgreSQL para poder pausar y
continuar en cualquier momento.

Diseñado para correr en tu computadora o en un servidor propio.
NO está pensado para GitHub Actions (sin límite de tiempo).

Flujo por ejecución:
  1. Leer checkpoint desde la DB (en qué página quedó)
  2. Procesar hasta MAX_PAGINAS_POR_RUN páginas desde ese punto
  3. Guardar checkpoint actualizado en la DB
  4. Salir limpiamente

Si el portal falla → guarda el checkpoint y sale sin error.
Cuando llega al final de todas las páginas → resetea a página 4
(las páginas 1-3 las cubre Fase 1 todos los días).

Uso:
  python fase2_scraper.py                  # Normal, usa checkpoint de la DB
  python fase2_scraper.py --max-paginas 5  # Útil para probar
  python fase2_scraper.py --reset          # Reinicia desde página 4
  python fase2_scraper.py --pagina 150     # Salta directo a una página
"""

import argparse
import asyncio
import json
import os
import re
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from playwright.async_api import async_playwright, Page

from sync_engine import crear_tablas, sync_proyectos, leer_checkpoint_db, guardar_checkpoint_db

# ----------------------------------------------------------------------
# CONFIGURACIÓN
# ----------------------------------------------------------------------

URL_BASE = (
    "https://www.asamblea.go.cr/Centro_de_informacion/"
    "Consultas_SIL/SitePages/SIL.aspx"
)

TEXTO_BOTON_ENTRADA  = "Expedientes Legislativos - Consulta"
MAX_PAGINAS_POR_RUN  = 5    # Páginas a procesar por ejecución
REGISTROS_POR_PAG    = "10"   # Dropdown de la grilla
PAGINA_INICIO_FASE2  = 11     # Fase 1 cubre 1-3, Fase 2 empieza en 4
MAX_PAGINA_TOTAL     = 2200   # Estimado de páginas totales (21.000 exp / 10)
MAX_REINTENTOS_SESION = 3     # Reintentos si la sesión se cae mid-run

# Tiempos de espera (ms)
ESPERA_CARGA_GRILLA  = 5_000
ESPERA_CLIC_FILA     = 2_500
ESPERA_CLIC_TAB      = 1_500
ESPERA_CLIC_PAGINA   = 4_000

IS_CI = os.getenv("CI", "false").lower() == "true"


# ----------------------------------------------------------------------
# ARGUMENTOS CLI
# ----------------------------------------------------------------------

def parsear_args():
    parser = argparse.ArgumentParser(
        description="Fase 2 - Backfill SIL Asamblea Legislativa"
    )
    parser.add_argument(
        "--max-paginas", type=int, default=None,
        help=f"Límite de páginas por ejecución (default: {MAX_PAGINAS_POR_RUN})"
    )
    parser.add_argument(
        "--reset", action="store_true",
        help=f"Reiniciar checkpoint a página {PAGINA_INICIO_FASE2}"
    )
    parser.add_argument(
        "--pagina", type=int, default=None,
        help="Saltar directamente a una página específica (ignora checkpoint)"
    )
    return parser.parse_args()


# ----------------------------------------------------------------------
# UTILIDADES
# ----------------------------------------------------------------------

def limpiar(v):
    """Elimina caracteres de control de una cadena."""
    if not isinstance(v, str):
        return v
    v = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', v)
    v = re.sub(r'[\u2400-\u243F]', '-', v)
    return v.strip()


def log(msg: str):
    """Print con timestamp."""
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}".encode('ascii', 'ignore').decode('ascii'), flush=True)


# ----------------------------------------------------------------------
# NAVEGACIÓN
# ----------------------------------------------------------------------

async def esperar_frame_webpart(page: Page, timeout_ms: int = 60_000) -> bool:
    inicio = asyncio.get_event_loop().time()
    while (asyncio.get_event_loop().time() - inicio) * 1000 < timeout_ms:
        for frame in page.frames:
            nombre = getattr(frame, 'name', '')
            if 'MSOPageViewer' in nombre or 'WebPartWPQ' in nombre:
                try:
                    count = await frame.evaluate(
                        "document.querySelectorAll('a, button').length"
                    )
                    if count > 0:
                        log(f"Frame '{nombre}' listo ({count} elementos).")
                        return True
                except Exception:
                    pass
        await page.wait_for_timeout(1_000)
    return False


async def navegar_a_expedientes(page: Page) -> bool:
    """
    Navega al módulo de expedientes. 5 intentos con 60s entre cada uno.
    """
    for intento in range(5):
        if intento > 0:
            log(f"Reintento {intento}/4 - recargando...")
            try:
                await page.goto(URL_BASE, wait_until="networkidle", timeout=60_000)
            except Exception:
                await page.goto(URL_BASE, wait_until="domcontentloaded", timeout=30_000)

        log("Esperando frame del portal...")
        if not await esperar_frame_webpart(page, timeout_ms=60_000):
            log("Frame no cargó. Esperando 60s...")
            await page.screenshot(path=f"debug_fase2_intento_{intento+1}.png", full_page=True)
            await page.wait_for_timeout(60_000)
            continue

        log(f"Buscando botón de entrada...")
        contextos = [page] + list(page.frames)

        for ctx in contextos:
            ctx_name = getattr(ctx, 'name', 'main')
            try:
                for sel in ["a[role='button']", "button", "a", "div[role='button']"]:
                    for boton in await ctx.query_selector_all(sel):
                        try:
                            texto = (await boton.inner_text()).strip()
                            if not texto:
                                texto = await boton.get_attribute("title") or ""
                            if not texto:
                                texto = await boton.get_attribute("aria-label") or ""
                            if TEXTO_BOTON_ENTRADA.lower() in texto.lower():
                                log(f"Botón encontrado en '{ctx_name}'. Haciendo clic...")
                                await boton.scroll_into_view_if_needed()
                                await boton.click()
                                await page.wait_for_load_state("networkidle", timeout=45_000)
                                await page.wait_for_timeout(ESPERA_CARGA_GRILLA)
                                return True
                        except Exception:
                            continue
            except Exception:
                continue

        # Fallback JS
        for ctx in contextos:
            try:
                clicked = await ctx.evaluate(f"""() => {{
                    const t = "{TEXTO_BOTON_ENTRADA}".toLowerCase();
                    for (const sel of ['a','button','div[role="button"]']) {{
                        for (const el of document.querySelectorAll(sel)) {{
                            const txt = (el.innerText || el.textContent ||
                                         el.getAttribute('title') ||
                                         el.getAttribute('aria-label') || '').toLowerCase();
                            if (txt.includes(t)) {{ el.click(); return true; }}
                        }}
                    }}
                    return false;
                }}""")
                if clicked:
                    await page.wait_for_load_state("networkidle", timeout=45_000)
                    await page.wait_for_timeout(ESPERA_CARGA_GRILLA)
                    return True
            except Exception:
                continue

        await page.screenshot(path=f"debug_fase2_intento_{intento+1}.png", full_page=True)
        await page.wait_for_timeout(60_000)

    return False


async def encontrar_frame_con_grilla(page: Page):
    for ctx in [page] + list(page.frames):
        try:
            for sel in ["div[role='grid']", ".jqx-grid", "div.jqx-grid-cell"]:
                if await ctx.query_selector(sel):
                    return ctx
        except Exception:
            continue
    return None


async def cambiar_registros_por_pagina(frame, page: Page, cantidad: str = "10") -> bool:
    try:
        for d in await frame.query_selector_all(".jqx-dropdownlist-content"):
            if (await d.inner_text()).strip().isdigit():
                await d.click()
                await page.wait_for_timeout(800)
                for ctx in [frame, page]:
                    opcion = await ctx.query_selector(
                        f".jqx-listitem-element:has-text('{cantidad}')"
                    )
                    if opcion:
                        await opcion.click()
                        await page.wait_for_timeout(ESPERA_CARGA_GRILLA)
                        return True
    except Exception as e:
        log(f"Error configurando registros: {e}")
    return False


async def obtener_rowscount(frame) -> int:
    try:
        return await frame.evaluate("""() => {
            const contenedor = document.querySelector('.marco-subcontenedor.alto-completo');
            if (!contenedor) return 0;
            const panel = [...contenedor.querySelectorAll('.jqx-tabs-content-element')].find(p => p.offsetParent !== null);
            if (!panel) return 0;
            const grilla = panel.querySelector("div[role='grid']");
            if (!grilla) return 0;
            const $ = window.$ || window.jQuery;
            if (!$ || !$(grilla).jqxGrid) return 0;
            try {
                return $(grilla).jqxGrid('getdatainformation').rowscount || 0;
            } catch(e) { return 0; }
        }""")
    except Exception:
        return 0


async def asegurar_50_registros(frame, page: Page, tab_name: str, rowscount: int, row_index: int):
    """
    Amplía la grilla de la pestaña actual a 50 registros usando clics en el UI.
    Y SIEMPRE restaura el foco en el expediente original para seguridad.
    """
    try:
        selector_dropdown = (
            '.marco-subcontenedor.alto-completo '
            '.jqx-tabs-content-element:not([style*="display: none"]) '
            '.jqx-dropdownlist-content'
        )
        drp = await frame.query_selector(selector_dropdown)
        if not drp:
            return
        
        texto_actual = await drp.inner_text()
        if "50" in texto_actual:
            return
        
        log(f"  > Ampliando '{tab_name}' ({rowscount} registros)...")
        await drp.click()
        await page.wait_for_timeout(1000)
        
        # Intentamos localizar la opción '50' tanto en el frame como en la página principal
        found = False
        for ctx in [frame, page]:
            opciones = ctx.locator(".jqx-item:visible, .jqx-listitem-element:visible").filter(has_text="50")
            if await opciones.count() > 0:
                await opciones.last.click(force=True)
                found = True
                break
        
        if found:
            await page.wait_for_timeout(1500)
        else:
            log(f"  > No se encontró la opción '50' en menús visibles.")
            
        # RESTAURACIÓN: Volver a tocar el expediente (plan sugerido por usuario)
        # Esto asegura que seguimos en el proyecto correcto aunque la grilla principal haya saltado.
        log(f"  > Restaurando foco en expediente (fila {row_index})...")
        await clicar_fila(frame, page, row_index)
        await clic_tab(frame, page, tab_name)
        await page.wait_for_timeout(600)
            
    except Exception as e:
        log(f"  > Error ajustando a 50 en {tab_name}: {e}")


async def iniciar_sesion_completa(page: Page, pagina_destino: int):
    """
    Navega al portal desde cero y salta a pagina_destino.
    Retorna (frame, True) si éxito, (None, False) si falla.
    """
    try:
        await page.goto(URL_BASE, wait_until="domcontentloaded", timeout=30_000)
    except Exception as e:
        log(f"Advertencia en recarga: {e}")

    await page.wait_for_timeout(5_000)

    if not await navegar_a_expedientes(page):
        return None, False

    await page.wait_for_timeout(3_000)
    frame = await encontrar_frame_con_grilla(page)
    if not frame:
        return None, False

    await cambiar_registros_por_pagina(frame, page, REGISTROS_POR_PAG)
    await page.wait_for_timeout(4_000)

    if pagina_destino > 1:
        log(f"Saltando a página {pagina_destino}...")
        if not await ir_a_pagina_directa(frame, page, pagina_destino):
            return None, False

    return frame, True


# ----------------------------------------------------------------------
# MODALES DE ERROR
# ----------------------------------------------------------------------

async def cerrar_modal_error(frame, page: Page) -> bool:
    try:
        modal = await frame.query_selector(".jqx-window-modal")
        if not modal or not await modal.is_visible():
            return False
        log("Modal de error detectado. Cerrando...")
        for sel in [
            ".jqx-window-close-button",
            "button:has-text('Cerrar')",
            "button:has-text('Aceptar')",
            "button:has-text('OK')",
        ]:
            try:
                btn = await frame.query_selector(sel)
                if btn and await btn.is_visible():
                    await btn.click()
                    await page.wait_for_timeout(600)
                    return True
            except Exception:
                continue
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(600)
        return True
    except Exception:
        return False


# ----------------------------------------------------------------------
# LECTURA DE FILAS
# ----------------------------------------------------------------------

async def obtener_info_filas(frame) -> list:
    info = []
    try:
        grilla = await frame.query_selector("div[role='grid']")
        if not grilla:
            return info
        for idx, row in enumerate(await grilla.query_selector_all("div[role='row']")):
            try:
                celdas = await row.query_selector_all("div[role='gridcell']")
                if len(celdas) < 2:
                    continue
                num    = (await celdas[0].inner_text()).strip()
                titulo = (await celdas[1].inner_text()).strip()
                if num and re.match(r'^\d{4,6}$', num.replace(" ", "")):
                    info.append({"expediente": num, "titulo": titulo, "row_index": idx})
            except Exception:
                continue
    except Exception as exc:
        log(f"Error leyendo grilla: {exc}")
    return info


async def clicar_fila(frame, page: Page, row_index: int) -> bool:
    await cerrar_modal_error(frame, page)
    try:
        rows = await frame.query_selector_all("div[role='row']")
        if row_index >= len(rows):
            return False
        await rows[row_index].click()
        await page.wait_for_timeout(ESPERA_CLIC_FILA)
        if await cerrar_modal_error(frame, page):
            return False
        return True
    except Exception as exc:
        log(f"  Error clic fila {row_index}: {exc}")
        await cerrar_modal_error(frame, page)
        return False


# ----------------------------------------------------------------------
# TABS DE DETALLE
# ----------------------------------------------------------------------

async def clic_tab(frame, page: Page, texto: str) -> bool:
    for sel in [
        f"td:has-text('{texto}')",
        f"li[role='tab']:has-text('{texto}')",
        f"div[role='tab']:has-text('{texto}')",
        f"a:has-text('{texto}')",
        f"span:has-text('{texto}')",
    ]:
        try:
            for el in await frame.query_selector_all(sel):
                if await el.is_visible():
                    await el.click()
                    await page.wait_for_timeout(ESPERA_CLIC_TAB)
                    return True
        except Exception:
            continue
    return False


async def extraer_tab_general(frame, page: Page) -> dict:
    await clic_tab(frame, page, "General")
    await page.wait_for_timeout(400)
    datos = {}
    try:
        datos = await frame.evaluate("""() => {
            const resultado = {};
            for (const tabla of document.querySelectorAll('table')) {
                if (tabla.closest("[role='grid']") || tabla.closest('.jqx-grid')) continue;
                for (const tr of tabla.querySelectorAll('tr')) {
                    const tds = [...tr.querySelectorAll('td')];
                    for (let i = 0; i < tds.length - 1; i++) {
                        if (tds[i].querySelector('input, select')) continue;
                        const label = (tds[i].innerText || tds[i].textContent || '')
                            .trim().replace(/[:：]$/, '').trim();
                        if (!label || label.length > 60) continue;
                        const inp = tds[i+1].querySelector('input, select');
                        const val = inp
                            ? (inp.value || inp.getAttribute('value') || '').trim()
                            : (tds[i+1].innerText || tds[i+1].textContent || '').trim();
                        if (label && val) resultado[label] = val;
                    }
                }
            }
            if (Object.keys(resultado).length === 0) {
                for (const inp of document.querySelectorAll('input[aria-label]')) {
                    const label = (inp.getAttribute('aria-label') || '').trim();
                    const val   = (inp.value || '').trim();
                    if (label && val) resultado[label] = val;
                }
            }
            return resultado;
        }""")
    except Exception as exc:
        log(f"  Error tab General: {exc}")
    return {k: limpiar(v) for k, v in datos.items()}


async def extraer_tab_tramitacion(frame, page: Page, row_index: int) -> list:
    await clic_tab(frame, page, "Tramitación")
    await page.wait_for_timeout(700)
    
    tramitacion = []
    try:
        # Extraer usando paginación si es necesario
        resultado = await extraer_grilla_paginada(
            frame, page, "Tramitación", row_index, 
            """() => {
                const contenedor = document.querySelector('.marco-subcontenedor.alto-completo');
                if (!contenedor) return null;
                const panel = [...contenedor.querySelectorAll('.jqx-tabs-content-element')]
                    .find(p => p.offsetParent !== null);
                if (!panel) return null;
                const grilla = panel.querySelector("div[role='grid']");
                if (!grilla) return null;
                const $ = window.$ || window.jQuery;
                if (!$ || !$(grilla).jqxGrid) return null;
                
                // Obtenemos solo los registros que están actualmente visibles/cargados en la página
                const rows = $(grilla).jqxGrid('getrows'); 
                if (!rows || rows.length === 0) return [];
                
                return rows.map(r => ({
                    organo:        String(r.Nombre_Corto        ?? ''),
                    descripcion:   String(r.Descripcion_Tramite ?? ''),
                    fecha_inicio:  String(r.Fecha_Inicio        ?? '').split(' ')[0],
                    fecha_termino: String(r.Fecha_Termino       ?? '').split(' ')[0],
                }));
            }"""
        )
        if resultado:
            for f in resultado:
                tramitacion.append({
                    "Órgano":        limpiar(f["organo"]),
                    "Descripción":   limpiar(f["descripcion"]),
                    "Fecha Inicio":  limpiar(f["fecha_inicio"]),
                    "Fecha Término": limpiar(f["fecha_termino"]),
                })
    except Exception as e:
        log(f"  Error tab Tramitación: {e}")
    return tramitacion


async def extraer_tab_proponentes(frame, page: Page, row_index: int) -> list:
    await clic_tab(frame, page, "Proponentes")
    await page.wait_for_timeout(700)
    
    proponentes = []
    try:
        resultado = await extraer_grilla_paginada(
            frame, page, "Proponentes", row_index,
            """() => {
                const contenedor = document.querySelector('.marco-subcontenedor.alto-completo');
                if (!contenedor) return null;
                const panel = [...contenedor.querySelectorAll('.jqx-tabs-content-element')]
                    .find(p => p.offsetParent !== null);
                if (!panel) return null;
                const grilla = panel.querySelector("div[role='grid']");
                if (!grilla) return null;
                const $ = window.$ || window.jQuery;
                if (!$ || !$(grilla).jqxGrid) return null;
                
                const rows = $(grilla).jqxGrid('getrows');
                if (!rows || rows.length === 0) return [];
                
                return rows.map(r => ({
                    firma:          String(r.Secuencia_Firma ?? ''),
                    nombre:         String(r.Nombre         ?? ''),
                    administracion: String(r.Administracion ?? ''),
                }));
            }"""
        )
        if resultado:
            for f in resultado:
                proponentes.append({
                "Firma":          limpiar(f["firma"]),
                    "Nombre":         limpiar(f["nombre"]),
                    "Administración": limpiar(f["administracion"]),
                })
    except Exception as e:
        log(f"  Error tab Proponentes: {e}")
    return proponentes


async def extraer_grilla_paginada(frame, page, tab_name, row_index, extractor_js: str) -> list:
    """
    Función genérica para extraer datos de una grilla con paginación optimizada.
    """
    total_datos = []
    rc = await obtener_rowscount(frame)

    if rc > 10:
        await asegurar_50_registros(frame, page, tab_name, rc, row_index)
        rc = await obtener_rowscount(frame)
        
    intentos = 0
    max_pags = 10
    
    while intentos < max_pags:
        datos_pagina = await frame.evaluate(extractor_js)
        if not datos_pagina:
            break
            
        total_datos.extend(datos_pagina)
        log(f"    > {tab_name}: página {intentos + 1} extraída ({len(total_datos)}/{rc if rc < 1000000 else '?'})")
            
        # -- LÓGICA DE SALIDA --
        # Salir si ya capturamos todo o si la página no está llena
        if (0 < rc < 1000000 and len(total_datos) >= rc) or len(datos_pagina) < 50:
            break
            
        selector_sig = (
            '.marco-subcontenedor.alto-completo '
            '.jqx-tabs-content-element:not([style*="display: none"]) '
            '.glyphicon-forward[title="Página siguiente"]'
        )
        
        btn_sig = await frame.query_selector(selector_sig)
        if not btn_sig:
            break
            
        log(f"    > Cargando siguiente página...")
        await btn_sig.click()
        await page.wait_for_timeout(2500)
        intentos += 1
        
    return total_datos


# ----------------------------------------------------------------------
# PAGINACIÓN
# ----------------------------------------------------------------------

async def ir_a_pagina_directa(frame, page: Page, numero: int) -> bool:
    """Navega directo a una página escribiendo en el input de paginación."""
    try:
        inp = (
            await frame.query_selector("input.ctrl-tabla-ira") or
            await frame.query_selector("input[title='Página actual']")
        )
        if not inp:
            return False

        filas_antes = await obtener_info_filas(frame)
        exp_antes = filas_antes[0]["expediente"] if filas_antes else None

        await inp.click(click_count=3)
        await inp.fill(str(numero))
        await inp.press("Enter")
        await page.wait_for_timeout(ESPERA_CLIC_PAGINA)

        for _ in range(12):
            filas = await obtener_info_filas(frame)
            exp = filas[0]["expediente"] if filas else None
            if exp and exp != exp_antes:
                log(f"Página {numero} cargada.")
                return True
            await page.wait_for_timeout(700)

        log(f"La grilla no cambió al saltar a página {numero}.")
        return False
    except Exception as e:
        log(f"Error saltando a página {numero}: {e}")
        return False


async def ir_siguiente_pagina(frame, page: Page, pagina_actual: int) -> bool:
    """Avanza a la siguiente página. Retorna True si hubo cambio real."""
    filas_antes = await obtener_info_filas(frame)
    exp_antes = filas_antes[0]["expediente"] if filas_antes else None

    clic_ok = False
    for sel in [
        "div[title='Siguiente página']",
        "input[title='Siguiente página']",
        "div[title='Next Page']",
        ".jqx-icon-arrow-right:not(.jqx-icon-arrow-right-selected)",
    ]:
        try:
            for btn in await frame.query_selector_all(sel):
                if not await btn.is_visible():
                    continue
                clase    = (await btn.get_attribute("class") or "").lower()
                disabled = await btn.get_attribute("disabled")
                title    = (await btn.get_attribute("title") or "").lower()
                if disabled or "disabled" in clase:
                    continue
                if "ltima" in title or "last" in title:
                    continue
                await btn.click()
                clic_ok = True
                break
        except Exception:
            continue
        if clic_ok:
            break

    # Fallback: buscar por texto o título (igual que fase1 que sí funciona)
    if not clic_ok:
        try:
            for btn in await frame.query_selector_all("div, button, a, input[type='button']"):
                try:
                    texto = (await btn.inner_text()).strip()
                    title = (await btn.get_attribute("title") or "").lower()
                    clase = (await btn.get_attribute("class") or "").lower()
                    if (texto in [">", "»", "›"] or "siguiente" in title or "next" in title):
                        if "disabled" not in clase and not await btn.get_attribute("disabled"):
                            await btn.click()
                            clic_ok = True
                            break
                except Exception:
                    continue
        except Exception:
            pass

    if not clic_ok:
        log("Botón 'Siguiente página' no encontrado o deshabilitado.")
        return False

    await page.wait_for_timeout(ESPERA_CLIC_PAGINA)
    for _ in range(10):
        filas = await obtener_info_filas(frame)
        exp = filas[0]["expediente"] if filas else None
        if exp and exp != exp_antes:
            return True
        await page.wait_for_timeout(700)

    return False


# ----------------------------------------------------------------------
# PROCESAR UNA PÁGINA
# ----------------------------------------------------------------------

async def procesar_pagina(
    page: Page,
    frame,
    num_pagina: int,
    acumulado: list,
) -> str:
    """
    Procesa todas las filas de una página.

    Retorna:
      "ok"            → página procesada correctamente
      "error_sesion"  → la grilla se cayó mid-página (reintentar con nueva sesión)
      "error_portal"  → fallo irrecuperable (parar el run)
    """
    await clic_tab(frame, page, "General")
    await page.wait_for_timeout(400)

    filas = await obtener_info_filas(frame)
    if not filas:
        log(f"Página {num_pagina}: sin filas. Posible error de portal.")
        return "error_portal"

    total = len(filas)
    log(f"{'-'*55}")
    log(f"PÁGINA {num_pagina} - {total} expedientes")
    log(f"{'-'*55}")

    filas_exitosas = 0

    for i, info in enumerate(filas):
        exp      = info["expediente"]
        titulo   = info["titulo"]
        row_idx  = info["row_index"]
        titulo_c = limpiar(titulo)[:60] + ("…" if len(titulo) > 60 else "")

        log(f"[{i+1}/{total}] Exp. {exp}: {titulo_c}")

        ok = await clicar_fila(frame, page, row_idx)
        if not ok:
            # Si las primeras filas fallan → probablemente la sesión cayó
            if filas_exitosas == 0 and i >= 1:
                log("  Múltiples filas fallidas desde el inicio → error de sesión.")
                return "error_sesion"
            log(f"  Fila omitida.")
            continue

        filas_exitosas += 1

        general     = await extraer_tab_general(frame, page)
        tramitacion = await extraer_tab_tramitacion(frame, page, row_idx)
        proponentes = await extraer_tab_proponentes(frame, page, row_idx)

        # Si después de varias filas exitosas los tabs desaparecen → sesión caída
        if filas_exitosas > 2 and len(general) == 0 and len(tramitacion) == 0:
            log("  Tabs vacíos tras filas exitosas → sesión caída.")
            return "error_sesion"

        await clic_tab(frame, page, "General")
        await page.wait_for_timeout(200)

        acumulado.append({
            "pagina":            num_pagina,
            "numero_expediente": exp,
            "titulo":            limpiar(titulo),
            "general":           general,
            "tramitacion":       tramitacion,
            "proponentes":       proponentes,
        })

        log(
            f"  ✓ General({len(general)}) "
            f"Tramitación({len(tramitacion)}) "
            f"Proponentes({len(proponentes)})"
        )

    if filas_exitosas == 0:
        log(f"  Ninguna fila pudo procesarse → error de sesión.")
        return "error_sesion"

    return "ok"


# ----------------------------------------------------------------------
# EXPORTAR EXCEL
# ----------------------------------------------------------------------

def exportar_excel(proyectos: list, nombre: str):
    h_fill = PatternFill("solid", fgColor="1F4E79")
    h_font = Font(bold=True, color="FFFFFF")
    centro = Alignment(horizontal="center")

    def enc(ws, cols):
        for c, txt in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=txt)
            cell.font = h_font; cell.fill = h_fill; cell.alignment = centro

    def s(v):
        return limpiar(v) if isinstance(v, str) else v

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    enc(ws, ["Página", "Expediente", "Título", "Tipo expediente",
             "Fecha inicio", "Vencimiento cuatrienal", "Ley"])
    for r, p in enumerate(proyectos, 2):
        g = p.get("general", {})
        ws.cell(r, 1, p.get("pagina"))
        ws.cell(r, 2, s(p.get("numero_expediente", "")))
        ws.cell(r, 3, s(p.get("titulo", "")))
        ws.cell(r, 4, s(g.get("Tipo expediente", "")))
        ws.cell(r, 5, s(g.get("Fecha inicio", "")))
        ws.cell(r, 6, s(g.get("Vencimiento cuatrienal", "")))
        ws.cell(r, 7, s(g.get("Ley", "")))
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 70

    ws_g = wb.create_sheet("General")
    enc(ws_g, ["Expediente", "Campo", "Valor"])
    r = 2
    for p in proyectos:
        for campo, val in p.get("general", {}).items():
            ws_g.cell(r, 1, s(p.get("numero_expediente", ""))); ws_g.cell(r, 2, s(campo)); ws_g.cell(r, 3, s(val)); r += 1

    ws_t = wb.create_sheet("Tramitación")
    enc(ws_t, ["Expediente", "Órgano", "Descripción", "Fecha Inicio", "Fecha Término"])
    r = 2
    for p in proyectos:
        for t in p.get("tramitacion", []):
            ws_t.cell(r, 1, s(p.get("numero_expediente", ""))); ws_t.cell(r, 2, s(t.get("Órgano", "")))
            ws_t.cell(r, 3, s(t.get("Descripción", ""))); ws_t.cell(r, 4, s(t.get("Fecha Inicio", ""))); ws_t.cell(r, 5, s(t.get("Fecha Término", ""))); r += 1

    ws_p = wb.create_sheet("Proponentes")
    enc(ws_p, ["Expediente", "Firma", "Nombre", "Administración"])
    r = 2
    for p in proyectos:
        for prop in p.get("proponentes", []):
            ws_p.cell(r, 1, s(p.get("numero_expediente", ""))); ws_p.cell(r, 2, s(prop.get("Firma", "")))
            ws_p.cell(r, 3, s(prop.get("Nombre", ""))); ws_p.cell(r, 4, s(prop.get("Administración", ""))); r += 1

    wb.save(nombre)
    log(f"Excel guardado: {nombre}")


# ----------------------------------------------------------------------
# MAIN
# ----------------------------------------------------------------------

async def main():
    args = parsear_args()
    max_pags = args.max_paginas or MAX_PAGINAS_POR_RUN
    inicio   = datetime.now()

    # Determinar página de inicio
    if args.reset:
        pagina_inicio = PAGINA_INICIO_FASE2
        log(f"--reset: checkpoint reiniciado a página {PAGINA_INICIO_FASE2}.")
    elif args.pagina:
        pagina_inicio = args.pagina
        log(f"--pagina: saltando directamente a página {pagina_inicio}.")
    else:
        pagina_inicio = leer_checkpoint_db()
        log(f"Checkpoint leído de DB: página {pagina_inicio}.")

    pagina_fin = pagina_inicio + max_pags - 1

    log("=" * 55)
    log("FASE 2 - Backfill SIL - Asamblea Legislativa CR")
    log("=" * 55)
    log(f"Inicio:        {inicio:%Y-%m-%d %H:%M:%S}")
    log(f"Rango:         páginas {pagina_inicio} -> {pagina_fin}")
    log(f"Máx por run:   {max_pags} páginas")
    log("=" * 55)

    proyectos    = []
    pagina_actual = pagina_inicio
    ciclo_completo = False

    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=IS_CI,
                args=["--no-sandbox", "--disable-dev-shm-usage",
                      "--disable-blink-features=AutomationControlled"]
            )
            context = await browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1440, "height": 900},
            )
            page = await context.new_page()

            # Carga inicial
            log("Cargando portal SIL...")
            try:
                await page.goto(URL_BASE, wait_until="networkidle", timeout=60_000)
            except Exception as e:
                log(f"Advertencia carga inicial: {e}")
                await page.goto(URL_BASE, wait_until="domcontentloaded", timeout=30_000)

            await page.wait_for_timeout(8_000)

            if not await navegar_a_expedientes(page):
                log("No se pudo navegar al módulo. Guardando checkpoint sin cambios.")
                guardar_checkpoint_db(pagina_inicio)
                await browser.close()
                sys.exit(0)

            await page.wait_for_timeout(3_000)
            frame = await encontrar_frame_con_grilla(page)
            if not frame:
                log("Grilla no encontrada. Guardando checkpoint sin cambios.")
                guardar_checkpoint_db(pagina_inicio)
                await browser.close()
                sys.exit(0)

            await cambiar_registros_por_pagina(frame, page, REGISTROS_POR_PAG)
            await page.wait_for_timeout(4_000)

            # Saltar a la página de inicio si no es la 1
            if pagina_inicio > 1:
                if not await ir_a_pagina_directa(frame, page, pagina_inicio):
                    log(f"No se pudo llegar a página {pagina_inicio}. Guardando checkpoint.")
                    guardar_checkpoint_db(pagina_inicio)
                    await browser.close()
                    sys.exit(0)

            # -- Loop principal -----------------------------------------
            paginas_procesadas = 0

            while paginas_procesadas < max_pags:

                # ¿Llegamos al final del portal?
                if pagina_actual > MAX_PAGINA_TOTAL:
                    log(f"Límite total alcanzado ({MAX_PAGINA_TOTAL}). Reiniciando ciclo.")
                    ciclo_completo = True
                    break

                resultado = await procesar_pagina(page, frame, pagina_actual, proyectos)

                # -- Error de sesión → intentar reiniciar --------------
                if resultado == "error_sesion":
                    log(f"Error de sesión en página {pagina_actual}. Intentando reiniciar...")
                    reintento_ok = False
                    for intento in range(MAX_REINTENTOS_SESION):
                        log(f"  Reintento de sesión {intento+1}/{MAX_REINTENTOS_SESION}...")
                        await page.wait_for_timeout(5_000)
                        frame_nuevo, ok = await iniciar_sesion_completa(page, pagina_actual)
                        if ok:
                            frame = frame_nuevo
                            reintento_ok = True
                            log(f"  Sesión reiniciada. Reintentando página {pagina_actual}...")
                            break
                        await page.wait_for_timeout(15_000)

                    if not reintento_ok:
                        log(f"Portal no responde tras {MAX_REINTENTOS_SESION} reintentos.")
                        log(f"Guardando checkpoint en página {pagina_actual} para continuar mañana.")
                        break

                    continue  # Reintentar la misma página

                # -- Error de portal → parar sin avanzar checkpoint ----
                elif resultado == "error_portal":
                    log(f"Error de portal en página {pagina_actual}. Deteniendo run.")
                    break

                # -- Página OK -----------------------------------------
                paginas_procesadas += 1
                log(f"Progreso: {paginas_procesadas}/{max_pags} páginas este run.")

                if paginas_procesadas >= max_pags:
                    log(f"Límite de {max_pags} páginas alcanzado para este run.")
                    break

                # Avanzar a la siguiente página
                if not await ir_siguiente_pagina(frame, page, pagina_actual):
                    filas_actuales = await obtener_info_filas(frame)
                    if not filas_actuales:
                        log("Sin más páginas disponibles. Ciclo completo.")
                        ciclo_completo = True
                    else:
                        # La página actual ya fue procesada OK - el checkpoint
                        # debe apuntar a la SIGUIENTE para no repetir trabajo.
                        pagina_actual += 1
                        log(f"No se pudo avanzar de página. Checkpoint en página {pagina_actual}.")
                    break

                pagina_actual += 1

            await browser.close()

    except Exception as e:
        log(f"Error inesperado (cancelación o cierre manual): {e}")
        log(f"Rescatando {len(proyectos)} expedientes recopilados hasta la página {pagina_actual}.")
        # NO hacemos 'raise' aquí para que pueda sincronizar a la DB lo que ya se descargó.

    stats = {}
    # -- Sync y exportación (PRIMERO SINCRONIZAMOS) ---------------------
    if proyectos:
        log("-" * 55)
        log("SINCRONIZACIÓN CON BASE DE DATOS")
        log("-" * 55)
        try:
            crear_tablas()
            stats = sync_proyectos(proyectos)
            log(f"¡Sincronización exitosa de {len(proyectos)} proyectos rescatados!")
        except Exception as e:
            log(f"Error crítico al sincronizar la BD: {e}")
            sys.exit(1) # Si falla la bd, salir SIN mover el checkpoint
    else:
        log("No se extrajeron proyectos en este run.")

    # -- Guardar checkpoint (SEGUNDO, SOLO SI SYNC FUE EXITOSO) ---------
    if ciclo_completo:
        proxima = PAGINA_INICIO_FASE2
        log(f"Ciclo completo. Próxima ejecución empezará en página {proxima}.")
    else:
        proxima = pagina_actual
        log(f"El próximo run arrancará en la página {proxima}.")

    guardar_checkpoint_db(proxima)

    ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_file  = f"fase2_{ts}_p{pagina_inicio}-{pagina_actual}.json"
    excel_file = f"fase2_{ts}_p{pagina_inicio}-{pagina_actual}.xlsx"

    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(proyectos, f, ensure_ascii=False, indent=2)
    log(f"JSON guardado: {json_file}")

    exportar_excel(proyectos, excel_file)

    duracion = datetime.now() - inicio

    log("=" * 55)
    log("RESUMEN FASE 2")
    log("=" * 55)
    log(f"Rango procesado:    páginas {pagina_inicio}-{pagina_actual}")
    log(f"Páginas procesadas: {paginas_procesadas}")
    log(f"Proyectos extraídos:{len(proyectos)}")
    log(f"DB sincronizados:   {stats.get('actualizados', 0)}")
    log(f"DB errores:         {stats.get('errores', 0)}")
    log(f"Próximo inicio:     página {proxima}")
    log(f"Duración total:     {str(duracion).split('.')[0]}")
    log("=" * 55)


if __name__ == "__main__":
    asyncio.run(main())