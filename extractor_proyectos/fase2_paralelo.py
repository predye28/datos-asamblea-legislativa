"""
fase2_paralelo.py
──────────────────────────────────────────────────────────────────────
Versión PARALELA del scraper Fase 2.

Misma lógica que fase2_scraper.py, pero lanza N workers de Chromium
en paralelo, cada uno procesando un rango distinto de páginas.

Configuración (variables a modificar aquí arriba):
  TOTAL_PAGINAS_POR_RUN  → cuántas páginas en total quieres procesar
  N_WORKERS              → cuántos browsers en paralelo

Ejemplo:
  TOTAL_PAGINAS_POR_RUN = 1000
  N_WORKERS = 3
  → 3 browsers corren en paralelo, cada uno procesa ~333 páginas

Checkpoint:
  - Lee la página de inicio desde la DB (igual que fase2_scraper.py)
  - Divide el rango entre N workers
  - Al terminar, guarda el checkpoint al final del rango total

Uso:
  python fase2_paralelo.py               # normal
  python fase2_paralelo.py --reset       # reinicia checkpoint a página 4
  python fase2_paralelo.py --pagina 150  # salta directo a una página
──────────────────────────────────────────────────────────────────────
"""

import argparse
import asyncio
import logging
import os
import re
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from playwright.async_api import async_playwright, Page

from sync_engine import crear_tablas, sync_proyectos, leer_checkpoint_db, guardar_checkpoint_db

# ══════════════════════════════════════════════════════════════════════
# ⚙️  CONFIGURACIÓN — MODIFICA ESTAS VARIABLES
# ══════════════════════════════════════════════════════════════════════

TOTAL_PAGINAS_POR_RUN = 10   # Total de páginas a procesar en este run
N_WORKERS             = 3    # Número de browsers en paralelo

# ══════════════════════════════════════════════════════════════════════
# CONSTANTES  (no es necesario cambiar estas)
# ══════════════════════════════════════════════════════════════════════

URL_BASE = (
    "https://www.asamblea.go.cr/Centro_de_informacion/"
    "Consultas_SIL/SitePages/SIL.aspx"
)

TEXTO_BOTON_ENTRADA   = "Expedientes Legislativos - Consulta"
REGISTROS_POR_PAG     = "10"
PAGINA_INICIO_FASE2   = 11
MAX_REINTENTOS_SESION = 3
MAX_PAGINA_TOTAL      = 99999  # backstop de emergencia; la detección real es por wrap

SLEEP_BETWEEN_BATCHES = 60    # segundos entre lotes en modo daemon
SLEEP_AT_CYCLE_END    = 3600  # segundos de espera al completar un ciclo completo

# Tiempos de espera (ms)
ESPERA_CARGA_GRILLA = 5_000
ESPERA_CLIC_FILA    = 2_500
ESPERA_CLIC_TAB     = 1_500
ESPERA_CLIC_PAGINA  = 4_000

IS_CI = os.getenv("CI", "false").lower() == "true"

logging.basicConfig(
    level=logging.INFO,
    format="%(message)s",
    handlers=[logging.StreamHandler()],
)
logger = logging.getLogger("fase2")


# ══════════════════════════════════════════════════════════════════════
# ARGUMENTOS CLI
# ══════════════════════════════════════════════════════════════════════

def parsear_args():
    parser = argparse.ArgumentParser(
        description="Fase 2 PARALELA - Backfill SIL Asamblea Legislativa"
    )
    parser.add_argument("--reset",  action="store_true",
                        help=f"Reiniciar checkpoint a página {PAGINA_INICIO_FASE2}")
    parser.add_argument("--pagina", type=int, default=None,
                        help="Saltar directamente a una página específica")
    parser.add_argument("--workers", type=int, default=None,
                        help=f"Número de workers (default: {N_WORKERS})")
    parser.add_argument("--total",   type=int, default=None,
                        help=f"Total de páginas a procesar (default: {TOTAL_PAGINAS_POR_RUN})")
    parser.add_argument("--daemon",  action="store_true",
                        help="Correr en bucle infinito (modo Docker/servidor)")
    return parser.parse_args()


# ══════════════════════════════════════════════════════════════════════
# UTILIDADES
# ══════════════════════════════════════════════════════════════════════

def limpiar(v):
    if not isinstance(v, str):
        return v
    v = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', v)
    v = re.sub(r'[\u2400-\u243F]', '-', v)
    return v.strip()


# Etiquetas de worker con ancho fijo para que las columnas queden alineadas
_WORKER_LABELS = [
    "W0", "W1", "W2", "W3", "W4",
    "W5", "W6", "W7", "W8", "W9",
]

def log(msg: str, worker_id: int = -1, nivel: str = "INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    prefix = f"[{_WORKER_LABELS[worker_id]}]" if 0 <= worker_id < len(_WORKER_LABELS) else "[MAIN]"
    iconos = {"INFO": " ", "WARN": "*", "ERROR": "!", "OK": "+"}
    icono = iconos.get(nivel, " ")
    line = f"[{ts}]{prefix}{icono} {msg}".encode("ascii", "ignore").decode("ascii")
    if nivel == "ERROR":
        logger.error(line)
    elif nivel == "WARN":
        logger.warning(line)
    else:
        logger.info(line)


def dividir_rangos(inicio: int, total: int, n: int) -> list[tuple[int, int]]:
    """
    Divide un rango de páginas equitativamente entre N workers.
    Retorna lista de (inicio, fin_inclusive) para cada worker.
    """
    paginas_por_worker = total // n
    sobra = total % n
    rangos = []
    pagina_actual = inicio
    for i in range(n):
        extra = 1 if i < sobra else 0
        cant = paginas_por_worker + extra
        fin = pagina_actual + cant - 1
        rangos.append((pagina_actual, fin))
        pagina_actual = fin + 1
    return rangos


# ══════════════════════════════════════════════════════════════════════
# MONITOR DE WORKERS
# ══════════════════════════════════════════════════════════════════════

class WorkerMonitor:
    """
    Coordina el estado de los workers para que si uno falla (especialmente
    uno de páginas bajas), los demás se enteren y se detengan.
    """
    def __init__(self):
        self._min_interrupted_page = float('inf')
        self._lock = asyncio.Lock()

    async def reportar_interrupcion(self, pagina: int):
        """Registra que un worker se detuvo en cierta página."""
        async with self._lock:
            if pagina < self._min_interrupted_page:
                self._min_interrupted_page = pagina

    async def deberia_detenerse(self, pagina_actual: int) -> tuple[bool, int]:
        """Consulta si este worker debería detenerse por una interrupción previa."""
        async with self._lock:
            if pagina_actual > self._min_interrupted_page:
                return True, int(self._min_interrupted_page)
            return False, 0


# ══════════════════════════════════════════════════════════════════════
# NAVEGACIÓN (idéntico a fase2_scraper.py)
# ══════════════════════════════════════════════════════════════════════

async def esperar_frame_webpart(page: Page, timeout_ms: int = 60_000) -> bool:
    inicio = asyncio.get_event_loop().time()
    while (asyncio.get_event_loop().time() - inicio) * 1000 < timeout_ms:
        for frame in page.frames:
            nombre = getattr(frame, 'name', '')
            if 'MSOPageViewer' in nombre or 'WebPartWPQ' in nombre:
                try:
                    count = await frame.evaluate(
                        "document.querySelectorAll('a, button').length"
                    )
                    if count > 0:
                        return True
                except Exception:
                    pass
        await page.wait_for_timeout(1_000)
    return False


async def navegar_a_expedientes(page: Page, worker_id: int = -1) -> bool:
    for intento in range(5):
        if intento > 0:
            log(f"Reintento {intento}/4 - recargando...", worker_id)
            try:
                await page.goto(URL_BASE, wait_until="networkidle", timeout=60_000)
            except Exception:
                await page.goto(URL_BASE, wait_until="domcontentloaded", timeout=30_000)

        if not await esperar_frame_webpart(page, timeout_ms=60_000):
            log("Frame no cargó. Esperando 60s...", worker_id)
            await page.wait_for_timeout(60_000)
            continue

        contextos = [page] + list(page.frames)

        for ctx in contextos:
            ctx_name = getattr(ctx, 'name', 'main')
            try:
                for sel in ["a[role='button']", "button", "a", "div[role='button']"]:
                    for boton in await ctx.query_selector_all(sel):
                        try:
                            texto = (await boton.inner_text()).strip()
                            if not texto:
                                texto = await boton.get_attribute("title") or ""
                            if not texto:
                                texto = await boton.get_attribute("aria-label") or ""
                            if TEXTO_BOTON_ENTRADA.lower() in texto.lower():
                                log(f"Botón encontrado en '{ctx_name}'.", worker_id)
                                await boton.scroll_into_view_if_needed()
                                await boton.click()
                                await page.wait_for_load_state("networkidle", timeout=45_000)
                                await page.wait_for_timeout(ESPERA_CARGA_GRILLA)
                                return True
                        except Exception:
                            continue
            except Exception:
                continue

        # Fallback JS
        for ctx in contextos:
            try:
                clicked = await ctx.evaluate(f"""() => {{
                    const t = "{TEXTO_BOTON_ENTRADA}".toLowerCase();
                    for (const sel of ['a','button','div[role="button"]']) {{
                        for (const el of document.querySelectorAll(sel)) {{
                            const txt = (el.innerText || el.textContent ||
                                         el.getAttribute('title') ||
                                         el.getAttribute('aria-label') || '').toLowerCase();
                            if (txt.includes(t)) {{ el.click(); return true; }}
                        }}
                    }}
                    return false;
                }}""")
                if clicked:
                    await page.wait_for_load_state("networkidle", timeout=45_000)
                    await page.wait_for_timeout(ESPERA_CARGA_GRILLA)
                    return True
            except Exception:
                continue

        await page.wait_for_timeout(60_000)

    return False


async def encontrar_frame_con_grilla(page: Page):
    for ctx in [page] + list(page.frames):
        try:
            for sel in ["div[role='grid']", ".jqx-grid", "div.jqx-grid-cell"]:
                if await ctx.query_selector(sel):
                    return ctx
        except Exception:
            continue
    return None


async def cambiar_registros_por_pagina(frame, page: Page, cantidad: str = "10") -> bool:
    try:
        for d in await frame.query_selector_all(".jqx-dropdownlist-content"):
            if (await d.inner_text()).strip().isdigit():
                await d.click()
                await page.wait_for_timeout(800)
                for ctx in [frame, page]:
                    opcion = await ctx.query_selector(
                        f".jqx-listitem-element:has-text('{cantidad}')"
                    )
                    if opcion:
                        await opcion.click()
                        await page.wait_for_timeout(ESPERA_CARGA_GRILLA)
                        return True
    except Exception as e:
        pass
    return False


async def obtener_rowscount(frame) -> int:
    try:
        return await frame.evaluate("""() => {
            const contenedor = document.querySelector('.marco-subcontenedor.alto-completo');
            if (!contenedor) return 0;
            const panel = [...contenedor.querySelectorAll('.jqx-tabs-content-element')].find(p => p.offsetParent !== null);
            if (!panel) return 0;
            const grilla = panel.querySelector("div[role='grid']");
            if (!grilla) return 0;
            const $ = window.$ || window.jQuery;
            if (!$ || !$(grilla).jqxGrid) return 0;
            try {
                return $(grilla).jqxGrid('getdatainformation').rowscount || 0;
            } catch(e) { return 0; }
        }""")
    except Exception:
        return 0


async def iniciar_sesion_completa(page: Page, pagina_destino: int, worker_id: int = -1):
    try:
        await page.goto(URL_BASE, wait_until="domcontentloaded", timeout=30_000)
    except Exception as e:
        log(f"Advertencia en recarga: {e}", worker_id)

    await page.wait_for_timeout(5_000)

    if not await navegar_a_expedientes(page, worker_id):
        return None, False

    await page.wait_for_timeout(3_000)
    frame = await encontrar_frame_con_grilla(page)
    if not frame:
        return None, False

    await cambiar_registros_por_pagina(frame, page, REGISTROS_POR_PAG)
    await page.wait_for_timeout(4_000)

    if pagina_destino > 1:
        log(f"Saltando a página {pagina_destino}...", worker_id)
        if not await ir_a_pagina_directa(frame, page, pagina_destino, worker_id):
            return None, False

    return frame, True


# ══════════════════════════════════════════════════════════════════════
# MODALES DE ERROR
# ══════════════════════════════════════════════════════════════════════

async def cerrar_modal_error(frame, page: Page) -> bool:
    try:
        modal = await frame.query_selector(".jqx-window-modal")
        if not modal or not await modal.is_visible():
            return False
        for sel in [".jqx-window-close-button", "button:has-text('Cerrar')",
                    "button:has-text('Aceptar')", "button:has-text('OK')"]:
            try:
                btn = await frame.query_selector(sel)
                if btn and await btn.is_visible():
                    await btn.click()
                    await page.wait_for_timeout(600)
                    return True
            except Exception:
                continue
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(600)
        return True
    except Exception:
        return False


# ══════════════════════════════════════════════════════════════════════
# LECTURA DE FILAS
# ══════════════════════════════════════════════════════════════════════

async def obtener_info_filas(frame) -> list:
    info = []
    try:
        grilla = await frame.query_selector("div[role='grid']")
        if not grilla:
            return info
        for idx, row in enumerate(await grilla.query_selector_all("div[role='row']")):
            try:
                celdas = await row.query_selector_all("div[role='gridcell']")
                if len(celdas) < 2:
                    continue
                num    = (await celdas[0].inner_text()).strip()
                titulo = (await celdas[1].inner_text()).strip()
                if num and re.match(r'^\d{4,6}$', num.replace(" ", "")):
                    info.append({"expediente": num, "titulo": titulo, "row_index": idx})
            except Exception:
                continue
    except Exception as exc:
        pass
    return info


async def clicar_fila(frame, page: Page, row_index: int) -> bool:
    await cerrar_modal_error(frame, page)
    try:
        rows = await frame.query_selector_all("div[role='row']")
        if row_index >= len(rows):
            return False
        await rows[row_index].click()
        await page.wait_for_timeout(ESPERA_CLIC_FILA)
        if await cerrar_modal_error(frame, page):
            return False
        return True
    except Exception as exc:
        await cerrar_modal_error(frame, page)
        return False


# ══════════════════════════════════════════════════════════════════════
# TABS DE DETALLE
# ══════════════════════════════════════════════════════════════════════

async def clic_tab(frame, page: Page, texto: str) -> bool:
    for sel in [
        f"td:has-text('{texto}')",
        f"li[role='tab']:has-text('{texto}')",
        f"div[role='tab']:has-text('{texto}')",
        f"a:has-text('{texto}')",
        f"span:has-text('{texto}')",
    ]:
        try:
            for el in await frame.query_selector_all(sel):
                if await el.is_visible():
                    await el.click()
                    await page.wait_for_timeout(ESPERA_CLIC_TAB)
                    return True
        except Exception:
            continue
    return False


async def extraer_tab_general(frame, page: Page) -> dict:
    await clic_tab(frame, page, "General")
    await page.wait_for_timeout(400)
    datos = {}
    try:
        datos = await frame.evaluate("""() => {
            const resultado = {};
            for (const tabla of document.querySelectorAll('table')) {
                if (tabla.closest("[role='grid']") || tabla.closest('.jqx-grid')) continue;
                for (const tr of tabla.querySelectorAll('tr')) {
                    const tds = [...tr.querySelectorAll('td')];
                    for (let i = 0; i < tds.length - 1; i++) {
                        if (tds[i].querySelector('input, select')) continue;
                        const label = (tds[i].innerText || tds[i].textContent || '')
                            .trim().replace(/[:：]$/, '').trim();
                        if (!label || label.length > 60) continue;
                        const inp = tds[i+1].querySelector('input, select');
                        const val = inp
                            ? (inp.value || inp.getAttribute('value') || '').trim()
                            : (tds[i+1].innerText || tds[i+1].textContent || '').trim();
                        if (label && val) resultado[label] = val;
                    }
                }
            }
            if (Object.keys(resultado).length === 0) {
                for (const inp of document.querySelectorAll('input[aria-label]')) {
                    const label = (inp.getAttribute('aria-label') || '').trim();
                    const val   = (inp.value || '').trim();
                    if (label && val) resultado[label] = val;
                }
            }
            return resultado;
        }""")
    except Exception as exc:
        pass
    return {k: limpiar(v) for k, v in datos.items()}


async def asegurar_50_registros(frame, page: Page, tab_name: str, rowscount: int, row_index: int, worker_id: int = -1):
    try:
        selector_dropdown = (
            '.marco-subcontenedor.alto-completo '
            '.jqx-tabs-content-element:not([style*="display: none"]) '
            '.jqx-dropdownlist-content'
        )
        drp = await frame.query_selector(selector_dropdown)
        if not drp:
            return

        texto_actual = await drp.inner_text()
        if "50" in texto_actual:
            return

        log(f"  > Ampliando '{tab_name}' ({rowscount} registros)...", worker_id)
        await drp.click()
        await page.wait_for_timeout(1000)

        found = False
        for ctx in [frame, page]:
            opciones = ctx.locator(".jqx-item:visible, .jqx-listitem-element:visible").filter(has_text="50")
            if await opciones.count() > 0:
                await opciones.last.click(force=True)
                found = True
                break

        if found:
            await page.wait_for_timeout(1500)

        log(f"  > Restaurando foco en expediente (fila {row_index})...", worker_id)
        await clicar_fila(frame, page, row_index)
        await clic_tab(frame, page, tab_name)
        await page.wait_for_timeout(600)

    except Exception as e:
        log(f"  > Error ajustando a 50 en {tab_name}: {e}", worker_id)


async def extraer_grilla_paginada(frame, page, tab_name, row_index, extractor_js: str, worker_id: int = -1) -> list:
    total_datos = []
    rc = await obtener_rowscount(frame)

    if rc > 10:
        await asegurar_50_registros(frame, page, tab_name, rc, row_index, worker_id)
        rc = await obtener_rowscount(frame)

    intentos = 0
    max_pags = 10

    while intentos < max_pags:
        datos_pagina = await frame.evaluate(extractor_js)
        if not datos_pagina:
            break

        total_datos.extend(datos_pagina)
        log(f"    > {tab_name}: pág {intentos + 1} extraída ({len(total_datos)}/{rc if rc < 1000000 else '?'})", worker_id)

        if (0 < rc < 1000000 and len(total_datos) >= rc) or len(datos_pagina) < 50:
            break

        selector_sig = (
            '.marco-subcontenedor.alto-completo '
            '.jqx-tabs-content-element:not([style*="display: none"]) '
            '.glyphicon-forward[title="Página siguiente"]'
        )

        btn_sig = await frame.query_selector(selector_sig)
        if not btn_sig:
            break

        log(f"    > Cargando siguiente página...", worker_id)
        await btn_sig.click()
        await page.wait_for_timeout(2500)
        intentos += 1

    return total_datos


async def extraer_tab_tramitacion(frame, page: Page, row_index: int, worker_id: int = -1) -> list:
    await clic_tab(frame, page, "Tramitación")
    await page.wait_for_timeout(700)

    tramitacion = []
    try:
        resultado = await extraer_grilla_paginada(
            frame, page, "Tramitación", row_index,
            """() => {
                const contenedor = document.querySelector('.marco-subcontenedor.alto-completo');
                if (!contenedor) return null;
                const panel = [...contenedor.querySelectorAll('.jqx-tabs-content-element')]
                    .find(p => p.offsetParent !== null);
                if (!panel) return null;
                const grilla = panel.querySelector("div[role='grid']");
                if (!grilla) return null;
                const $ = window.$ || window.jQuery;
                if (!$ || !$(grilla).jqxGrid) return null;
                const rows = $(grilla).jqxGrid('getrows');
                if (!rows || rows.length === 0) return [];
                return rows.map(r => ({
                    organo:        String(r.Nombre_Corto        ?? ''),
                    descripcion:   String(r.Descripcion_Tramite ?? ''),
                    fecha_inicio:  String(r.Fecha_Inicio        ?? '').split(' ')[0],
                    fecha_termino: String(r.Fecha_Termino       ?? '').split(' ')[0],
                }));
            }""",
            worker_id
        )
        if resultado:
            for f in resultado:
                tramitacion.append({
                    "Órgano":        limpiar(f["organo"]),
                    "Descripción":   limpiar(f["descripcion"]),
                    "Fecha Inicio":  limpiar(f["fecha_inicio"]),
                    "Fecha Término": limpiar(f["fecha_termino"]),
                })
    except Exception as e:
        log(f"  Error tab Tramitación: {e}", worker_id)
    return tramitacion


async def extraer_tab_proponentes(frame, page: Page, row_index: int, worker_id: int = -1) -> list:
    await clic_tab(frame, page, "Proponentes")
    await page.wait_for_timeout(700)

    proponentes = []
    try:
        resultado = await extraer_grilla_paginada(
            frame, page, "Proponentes", row_index,
            """() => {
                const contenedor = document.querySelector('.marco-subcontenedor.alto-completo');
                if (!contenedor) return null;
                const panel = [...contenedor.querySelectorAll('.jqx-tabs-content-element')]
                    .find(p => p.offsetParent !== null);
                if (!panel) return null;
                const grilla = panel.querySelector("div[role='grid']");
                if (!grilla) return null;
                const $ = window.$ || window.jQuery;
                if (!$ || !$(grilla).jqxGrid) return null;
                const rows = $(grilla).jqxGrid('getrows');
                if (!rows || rows.length === 0) return [];
                return rows.map(r => ({
                    firma:          String(r.Secuencia_Firma ?? ''),
                    nombre:         String(r.Nombre         ?? ''),
                    administracion: String(r.Administracion ?? ''),
                }));
            }""",
            worker_id
        )
        if resultado:
            for f in resultado:
                proponentes.append({
                    "Firma":          limpiar(f["firma"]),
                    "Nombre":         limpiar(f["nombre"]),
                    "Administración": limpiar(f["administracion"]),
                })
    except Exception as e:
        log(f"  Error tab Proponentes: {e}", worker_id)
    return proponentes


# ══════════════════════════════════════════════════════════════════════
# PAGINACIÓN
# ══════════════════════════════════════════════════════════════════════

async def ir_a_pagina_directa(frame, page: Page, numero: int, worker_id: int = -1) -> bool:
    try:
        inp = (
            await frame.query_selector("input.ctrl-tabla-ira") or
            await frame.query_selector("input[title='Página actual']")
        )
        if not inp:
            return False

        filas_antes = await obtener_info_filas(frame)
        exp_antes = filas_antes[0]["expediente"] if filas_antes else None

        await inp.click(click_count=3)
        await inp.fill(str(numero))
        await inp.press("Enter")
        await page.wait_for_timeout(ESPERA_CLIC_PAGINA)

        for _ in range(12):
            filas = await obtener_info_filas(frame)
            exp = filas[0]["expediente"] if filas else None
            if exp and exp != exp_antes:
                log(f"Página {numero} cargada.", worker_id)
                return True
            await page.wait_for_timeout(700)

        log(f"La grilla no cambió al saltar a página {numero}.", worker_id)
        return False
    except Exception as e:
        log(f"Error saltando a página {numero}: {e}", worker_id)
        return False


async def ir_siguiente_pagina(frame, page: Page, pagina_actual: int) -> bool:
    filas_antes = await obtener_info_filas(frame)
    exp_antes = filas_antes[0]["expediente"] if filas_antes else None

    clic_ok = False
    for sel in [
        "div[title='Siguiente página']",
        "input[title='Siguiente página']",
        "div[title='Next Page']",
        ".jqx-icon-arrow-right:not(.jqx-icon-arrow-right-selected)",
    ]:
        try:
            for btn in await frame.query_selector_all(sel):
                if not await btn.is_visible():
                    continue
                clase    = (await btn.get_attribute("class") or "").lower()
                disabled = await btn.get_attribute("disabled")
                title    = (await btn.get_attribute("title") or "").lower()
                if disabled or "disabled" in clase:
                    continue
                if "ltima" in title or "last" in title:
                    continue
                await btn.click()
                clic_ok = True
                break
        except Exception:
            continue
        if clic_ok:
            break

    if not clic_ok:
        try:
            for btn in await frame.query_selector_all("div, button, a, input[type='button']"):
                try:
                    texto = (await btn.inner_text()).strip()
                    title = (await btn.get_attribute("title") or "").lower()
                    clase = (await btn.get_attribute("class") or "").lower()
                    if (texto in [">", "»", "›"] or "siguiente" in title or "next" in title):
                        if "disabled" not in clase and not await btn.get_attribute("disabled"):
                            await btn.click()
                            clic_ok = True
                            break
                except Exception:
                    continue
        except Exception:
            pass

    if not clic_ok:
        return False

    await page.wait_for_timeout(ESPERA_CLIC_PAGINA)
    for _ in range(10):
        filas = await obtener_info_filas(frame)
        exp = filas[0]["expediente"] if filas else None
        if exp and exp != exp_antes:
            return True
        await page.wait_for_timeout(700)

    return False


async def leer_pagina_actual(frame) -> int | None:
    """Lee el número de página real que muestra el sitio en el input de paginación."""
    try:
        inp = (
            await frame.query_selector("input.ctrl-tabla-ira") or
            await frame.query_selector("input[title='Página actual']")
        )
        if not inp:
            return None
        val = await inp.input_value()
        return int(val) if val and val.strip().isdigit() else None
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════
# EXPORTAR EXCEL
# ══════════════════════════════════════════════════════════════════════

def exportar_excel(proyectos: list, nombre: str):
    h_fill = PatternFill("solid", fgColor="1F4E79")
    h_font = Font(bold=True, color="FFFFFF")
    centro = Alignment(horizontal="center")

    def enc(ws, cols):
        for c, txt in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=txt)
            cell.font = h_font; cell.fill = h_fill; cell.alignment = centro

    def s(v):
        return limpiar(v) if isinstance(v, str) else v

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    enc(ws, ["Página", "Expediente", "Título", "Tipo expediente",
             "Fecha inicio", "Vencimiento cuatrienal", "Ley"])
    for r, p in enumerate(proyectos, 2):
        g = p.get("general", {})
        ws.cell(r, 1, p.get("pagina"))
        ws.cell(r, 2, s(p.get("numero_expediente", "")))
        ws.cell(r, 3, s(p.get("titulo", "")))
        ws.cell(r, 4, s(g.get("Tipo expediente", "")))
        ws.cell(r, 5, s(g.get("Fecha inicio", "")))
        ws.cell(r, 6, s(g.get("Vencimiento cuatrienal", "")))
        ws.cell(r, 7, s(g.get("Ley", "")))
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 70

    ws_t = wb.create_sheet("Tramitación")
    enc(ws_t, ["Expediente", "Órgano", "Descripción", "Fecha Inicio", "Fecha Término"])
    r = 2
    for p in proyectos:
        for t in p.get("tramitacion", []):
            ws_t.cell(r, 1, s(p.get("numero_expediente", ""))); ws_t.cell(r, 2, s(t.get("Órgano", "")))
            ws_t.cell(r, 3, s(t.get("Descripción", ""))); ws_t.cell(r, 4, s(t.get("Fecha Inicio", ""))); ws_t.cell(r, 5, s(t.get("Fecha Término", ""))); r += 1

    ws_p = wb.create_sheet("Proponentes")
    enc(ws_p, ["Expediente", "Firma", "Nombre", "Administración"])
    r = 2
    for p in proyectos:
        for prop in p.get("proponentes", []):
            ws_p.cell(r, 1, s(p.get("numero_expediente", ""))); ws_p.cell(r, 2, s(prop.get("Firma", "")))
            ws_p.cell(r, 3, s(prop.get("Nombre", ""))); ws_p.cell(r, 4, s(prop.get("Administración", ""))); r += 1

    wb.save(nombre)
    log(f"Excel guardado: {nombre}")


# ══════════════════════════════════════════════════════════════════════
# PROCESAR UNA PÁGINA (idéntico a fase2_scraper.py)
# ══════════════════════════════════════════════════════════════════════

async def procesar_pagina(page: Page, frame, num_pagina: int, acumulado: list, worker_id: int = -1) -> str:
    await clic_tab(frame, page, "General")
    await page.wait_for_timeout(400)

    filas = await obtener_info_filas(frame)
    if not filas:
        log(f"Página {num_pagina}: sin filas. Posible error de portal.", worker_id)
        return "error_portal"

    total = len(filas)
    log(f"{'—'*50}", worker_id)
    log(f"PÁGINA {num_pagina} — {total} expedientes", worker_id)
    log(f"{'—'*50}", worker_id)

    filas_exitosas = 0

    for i, info in enumerate(filas):
        exp      = info["expediente"]
        titulo   = info["titulo"]
        row_idx  = info["row_index"]
        titulo_c = limpiar(titulo)[:55] + ("…" if len(titulo) > 55 else "")

        log(f"[{i+1}/{total}] Exp. {exp}: {titulo_c}", worker_id)

        ok = await clicar_fila(frame, page, row_idx)
        if not ok:
            if filas_exitosas == 0 and i >= 1:
                log("  Múltiples filas fallidas → error de sesión.", worker_id)
                return "error_sesion"
            log(f"  Fila omitida.", worker_id)
            continue

        filas_exitosas += 1

        general     = await extraer_tab_general(frame, page)
        tramitacion = await extraer_tab_tramitacion(frame, page, row_idx, worker_id)
        proponentes = await extraer_tab_proponentes(frame, page, row_idx, worker_id)

        if filas_exitosas > 2 and len(general) == 0 and len(tramitacion) == 0:
            log("  Tabs vacíos tras filas exitosas → sesión caída.", worker_id)
            return "error_sesion"

        await clic_tab(frame, page, "General")
        await page.wait_for_timeout(200)

        acumulado.append({
            "pagina":            num_pagina,
            "numero_expediente": exp,
            "titulo":            limpiar(titulo),
            "general":           general,
            "tramitacion":       tramitacion,
            "proponentes":       proponentes,
        })

        log(
            f"  ✓ General({len(general)}) "
            f"Tramitación({len(tramitacion)}) "
            f"Proponentes({len(proponentes)})",
            worker_id
        )

    if filas_exitosas == 0:
        log(f"  Ninguna fila pudo procesarse → error de sesión.", worker_id)
        return "error_sesion"

    return "ok"


# ══════════════════════════════════════════════════════════════════════
# WORKER INDIVIDUAL — un browser completo para un rango de páginas
# ══════════════════════════════════════════════════════════════════════

async def worker(
    playwright,
    pagina_inicio: int,
    pagina_fin: int,
    worker_id: int,
    monitor: WorkerMonitor
) -> tuple[list, int, bool, bool]:
    """
    Lanza un browser independiente y procesa el rango [pagina_inicio, pagina_fin].

    Retorna:
      (proyectos_extraidos, ultima_pagina_procesada, completado_sin_errores, ciclo_completo)

    `ciclo_completo` es True cuando el sitio hizo wrap a la página 1, indicando
    que se recorrió toda la base de datos y hay que reiniciar desde el inicio.
    """
    proyectos      = []
    pagina_actual  = pagina_inicio
    completado     = False
    ciclo_completo = False

    log(f"Iniciando — rango paginas {pagina_inicio} -> {pagina_fin} ({pagina_fin - pagina_inicio + 1} pags)", worker_id)

    try:
        browser = await playwright.chromium.launch(
            headless=IS_CI,
            args=["--no-sandbox", "--disable-dev-shm-usage",
                  "--disable-blink-features=AutomationControlled"]
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1440, "height": 900},
        )
        page = await context.new_page()

        log("Cargando portal SIL...", worker_id)
        try:
            await page.goto(URL_BASE, wait_until="networkidle", timeout=60_000)
        except Exception as e:
            log(f"Advertencia carga inicial: {e}", worker_id, "WARN")
            await page.goto(URL_BASE, wait_until="domcontentloaded", timeout=30_000)

        await page.wait_for_timeout(8_000)

        if not await navegar_a_expedientes(page, worker_id):
            log("No se pudo navegar al modulo. Terminando worker.", worker_id, "ERROR")
            await browser.close()
            return proyectos, pagina_actual, completado, ciclo_completo

        await page.wait_for_timeout(3_000)
        frame = await encontrar_frame_con_grilla(page)
        if not frame:
            log("Grilla no encontrada. Terminando worker.", worker_id, "ERROR")
            await browser.close()
            return proyectos, pagina_actual, completado, ciclo_completo

        await cambiar_registros_por_pagina(frame, page, REGISTROS_POR_PAG)
        await page.wait_for_timeout(4_000)

        if pagina_inicio > 1:
            if not await ir_a_pagina_directa(frame, page, pagina_inicio, worker_id):
                log(f"No se pudo llegar a pagina {pagina_inicio}. Terminando worker.", worker_id, "ERROR")
                await browser.close()
                return proyectos, pagina_actual, completado, ciclo_completo

            # Si el sitio devolvió una página menor a la pedida, hizo wrap → fin del dataset
            pagina_real = await leer_pagina_actual(frame)
            log(f"Salto directo: pag real={pagina_real} | pedida={pagina_inicio}", worker_id)
            if pagina_real is not None and pagina_real < pagina_inicio:
                log(f"Wrap al saltar: sitio mostró pag {pagina_real} (pedida {pagina_inicio}). Ciclo completo.", worker_id, "OK")
                ciclo_completo = True
                completado = True
                await browser.close()
                return proyectos, pagina_actual, completado, ciclo_completo

        # ── Loop de páginas del worker ──────────────────────────────
        paginas_procesadas = 0
        max_paginas_worker = pagina_fin - pagina_inicio + 1

        while paginas_procesadas < max_paginas_worker:
            # --- Check de monitor ---
            stoppy, pag_fallida = await monitor.deberia_detenerse(pagina_actual)
            if stoppy:
                log(f"DETENCIÓN PREVENTIVA: Se detectó fallo en página {pag_fallida}. Cancelando trabajo redundante.", worker_id, "WARN")
                break

            if pagina_actual > MAX_PAGINA_TOTAL:
                log(f"Limite total ({MAX_PAGINA_TOTAL}) alcanzado.", worker_id, "WARN")
                break

            resultado = await procesar_pagina(page, frame, pagina_actual, proyectos, worker_id)

            if resultado == "error_sesion":
                log(f"Error de sesion en pagina {pagina_actual}. Intentando reinicio...", worker_id, "WARN")
                reintento_ok = False
                for intento in range(MAX_REINTENTOS_SESION):
                    log(f"  Reintento {intento+1}/{MAX_REINTENTOS_SESION}...", worker_id)
                    await page.wait_for_timeout(5_000)
                    frame_nuevo, ok = await iniciar_sesion_completa(page, pagina_actual, worker_id)
                    if ok:
                        frame = frame_nuevo
                        reintento_ok = True
                        log(f"  Sesion restablecida. Continuando desde pagina {pagina_actual}.", worker_id, "OK")
                        break
                    await page.wait_for_timeout(15_000)

                if not reintento_ok:
                    log(f"Portal no responde tras {MAX_REINTENTOS_SESION} reintentos. Worker detenido en pagina {pagina_actual}.", worker_id, "ERROR")
                    await monitor.reportar_interrupcion(pagina_actual)
                    break  # completado sigue False, pagina_actual = dónde se cayó
                continue

            elif resultado == "error_portal":
                log(f"Error de portal en pagina {pagina_actual}. Deteniendo worker.", worker_id, "ERROR")
                await monitor.reportar_interrupcion(pagina_actual)
                break

            # Página OK
            paginas_procesadas += 1
            pct = round(paginas_procesadas / max_paginas_worker * 100)
            log(
                f"Pagina {pagina_actual} OK | {paginas_procesadas}/{max_paginas_worker} ({pct}%)",
                worker_id, "OK"
            )

            if paginas_procesadas >= max_paginas_worker:
                completado = True  # termino su rango sin caerse
                break

            if not await ir_siguiente_pagina(frame, page, pagina_actual):
                log("No se pudo avanzar de pagina.", worker_id, "WARN")
                pagina_actual += 1
                break

            # Si el sitio volvió a una página anterior, detectamos el wrap (fin del dataset)
            pagina_real = await leer_pagina_actual(frame)
            log(f"Navegacion: pag real={pagina_real} | esperada={pagina_actual + 1}", worker_id)
            if pagina_real is not None and pagina_real < pagina_actual:
                log(f"Wrap detectado: sitio mostró pag {pagina_real} (esperada {pagina_actual + 1}). Ciclo completo.", worker_id, "OK")
                ciclo_completo = True
                completado = True
                pagina_actual += 1
                break

            pagina_actual += 1

        await browser.close()

    except Exception as e:
        log(f"Excepcion inesperada: {e}", worker_id, "ERROR")
        await monitor.reportar_interrupcion(pagina_actual)
        # pagina_actual queda donde estaba → el main lo usa para el checkpoint

    estado = "COMPLETO" if completado else f"INTERRUMPIDO en pag {pagina_actual}"
    log(f"Fin worker | {len(proyectos)} proyectos | {estado}", worker_id, "OK" if completado else "WARN")
    return proyectos, pagina_actual, completado, ciclo_completo


# ══════════════════════════════════════════════════════════════════════
# SEÑALES
# ══════════════════════════════════════════════════════════════════════

_stop = False

def _setup_signal_handlers():
    import signal
    def _handler(_sig, _frame):
        global _stop
        log("Senal de parada recibida. Terminando al final del lote actual...", nivel="WARN")
        _stop = True
    signal.signal(signal.SIGTERM, _handler)
    signal.signal(signal.SIGINT, _handler)

async def _interruptible_sleep(seconds: int):
    """Sleep que se interrumpe en menos de 1s si llega SIGTERM/SIGINT."""
    for _ in range(seconds):
        if _stop:
            break
        await asyncio.sleep(1)


# ══════════════════════════════════════════════════════════════════════
# MAIN PARALELO
# ══════════════════════════════════════════════════════════════════════

async def main():
    args = parsear_args()
    _setup_signal_handlers()

    n_workers  = args.workers or N_WORKERS
    total_pags = args.total   or TOTAL_PAGINAS_POR_RUN
    primera_iter = True

    while True:
        if _stop:
            log("Parada solicitada. Saliendo limpiamente.", nivel="WARN")
            break

        inicio = datetime.now()

        # --reset y --pagina solo aplican en la primera iteracion
        if primera_iter and args.reset:
            pagina_inicio = PAGINA_INICIO_FASE2
            log(f"--reset: checkpoint reiniciado a pagina {PAGINA_INICIO_FASE2}")
        elif primera_iter and args.pagina:
            pagina_inicio = args.pagina
            log(f"--pagina: saltando directamente a pagina {pagina_inicio}")
        else:
            pagina_inicio = leer_checkpoint_db()
            log(f"Checkpoint leido de DB: pagina {pagina_inicio}")
        primera_iter = False

        pagina_fin = pagina_inicio + total_pags - 1
        rangos = dividir_rangos(pagina_inicio, total_pags, n_workers)

        log("=" * 60)
        log("FASE 2 PARALELA — Backfill SIL Asamblea Legislativa CR")
        log("=" * 60)
        log(f"Inicio:          {inicio:%Y-%m-%d %H:%M:%S}")
        log(f"Rango total:     paginas {pagina_inicio} -> {pagina_fin}")
        log(f"Total paginas:   {total_pags}")
        log(f"Workers:         {n_workers}")
        for i, (ini, fin) in enumerate(rangos):
            log(f"  Worker {i}: paginas {ini} -> {fin} ({fin-ini+1} paginas)")
        log("=" * 60)
        log(f"Lanzando {n_workers} workers en paralelo...")
        log("=" * 60)

        monitor = WorkerMonitor()

        async with async_playwright() as playwright:
            tareas = [
                worker(playwright, ini, fin, worker_id=i, monitor=monitor)
                for i, (ini, fin) in enumerate(rangos)
            ]
            resultados = await asyncio.gather(*tareas, return_exceptions=True)

        # ── Analizar resultados de cada worker ─────────────────────────
        log("=" * 60)
        log("Resultados por worker:")
        log("=" * 60)

        todos_proyectos     = []
        workers_completos   = []
        workers_incompletos = []
        wrap_detectado      = False

        for i, resultado in enumerate(resultados):
            ini_w, fin_w = rangos[i]

            if isinstance(resultado, Exception):
                log(f"  Worker {i} FALLO con excepcion: {resultado}", nivel="ERROR")
                workers_incompletos.append((i, ini_w))
                continue

            proyectos_w, ultima_pag_w, completo_w, ciclo_w = resultado
            todos_proyectos.extend(proyectos_w)
            if ciclo_w:
                wrap_detectado = True

            if completo_w:
                log(f"  Worker {i} COMPLETO  | pags {ini_w}-{fin_w} | {len(proyectos_w)} proyectos", nivel="OK")
                workers_completos.append(i)
            else:
                log(f"  Worker {i} INCOMPLETO| pags {ini_w}-{fin_w} | cayo en pag {ultima_pag_w} | {len(proyectos_w)} proyectos", nivel="WARN")
                workers_incompletos.append((i, ultima_pag_w))

        # ── Determinar checkpoint seguro ───────────────────────────────
        if wrap_detectado:
            # El sitio llegó a la última página y volvió al inicio: ciclo completo
            proxima = PAGINA_INICIO_FASE2
            log("Wrap detectado: dataset recorrido completo. Proximo ciclo desde el inicio.", nivel="OK")
        elif workers_incompletos:
            pagina_minima_caida = min(pag for _, pag in workers_incompletos)
            proxima = pagina_minima_caida
            log("", nivel="WARN")
            log(f"ATENCION: {len(workers_incompletos)} worker(s) no completaron su rango.", nivel="WARN")
            log(f"  Workers incompletos: {[f'W{i}@pag{p}' for i, p in workers_incompletos]}", nivel="WARN")
            log(f"  Checkpoint guardado en pagina {proxima} (primera pagina con dato faltante).", nivel="WARN")
        else:
            proxima = pagina_fin + 1
            log(f"Todos los workers completaron. Proximo run desde pagina {proxima}.", nivel="OK")

        todos_proyectos.sort(key=lambda p: int(p.get("numero_expediente", 0) or 0))
        log(f"Total combinado: {len(todos_proyectos)} proyectos de {n_workers} workers.")

        # ── Sync a base de datos ───────────────────────────────────────
        stats = {}
        if todos_proyectos:
            log("-" * 55)
            log("SINCRONIZACION CON BASE DE DATOS")
            log("-" * 55)
            try:
                crear_tablas()
                stats = sync_proyectos(todos_proyectos)
                log(f"Sincronizacion exitosa de {len(todos_proyectos)} proyectos.", nivel="OK")
            except Exception as e:
                log(f"Error critico al sincronizar la BD: {e}", nivel="ERROR")
                guardar_checkpoint_db(proxima)
                if not args.daemon:
                    sys.exit(1)
                log("Modo daemon: continuando en el proximo lote...", nivel="WARN")
        else:
            log("No se extrajeron proyectos en este run.", nivel="WARN")

        # ── Guardar checkpoint ─────────────────────────────────────────
        guardar_checkpoint_db(proxima)

        # ── Heartbeat para Docker health check ────────────────────────
        try:
            with open("/tmp/scraper_heartbeat", "w") as hb:
                hb.write(datetime.now().isoformat())
        except Exception:
            pass

        duracion = datetime.now() - inicio

        log("=" * 60)
        log("RESUMEN FASE 2 PARALELA")
        log("=" * 60)
        log(f"Rango planificado: pags {pagina_inicio} -> {pagina_fin}")
        log(f"Workers usados:    {n_workers}")
        log(f"  Completados:     {len(workers_completos)}")
        log(f"  Incompletos:     {len(workers_incompletos)}")
        log(f"Proyectos totales: {len(todos_proyectos)}")
        log(f"DB sincronizados:  {stats.get('actualizados', 0)}")
        log(f"DB errores:        {stats.get('errores', 0)}")
        log(f"Proximo inicio:    pagina {proxima}")
        log(f"Duracion total:    {str(duracion).split('.')[0]}")
        log("=" * 60)

        # ── Modo single run: terminar ──────────────────────────────────
        if not args.daemon:
            break

        # ── Modo daemon: pausar antes del siguiente lote ───────────────
        if proxima == PAGINA_INICIO_FASE2:
            log(f"Esperando {SLEEP_AT_CYCLE_END}s antes del proximo ciclo...", nivel="INFO")
            await _interruptible_sleep(SLEEP_AT_CYCLE_END)
        else:
            log(f"Esperando {SLEEP_BETWEEN_BATCHES}s antes del siguiente lote...", nivel="INFO")
            await _interruptible_sleep(SLEEP_BETWEEN_BATCHES)


if __name__ == "__main__":
    asyncio.run(main())
