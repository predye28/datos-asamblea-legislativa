"""
extraer_proyectos.py
----------------------------------------------------------------------
Scraper for bill projects from the Legislative Assembly of Costa Rica.

Applied fixes:
  - Re-reads grid rows in each iteration to avoid stale element references.
  - Identifies the bottom panel by position to distinguish from the main grid.
  - Extracts processing steps and proponents from the correct containers.
"""

import argparse
import asyncio
import json
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from playwright.async_api import async_playwright, Page

from sync_engine import crear_tablas, sync_proyectos

# ----------------------------------------------------------------------
# CONFIGURATION
# ----------------------------------------------------------------------

URL_BASE = (
    "https://www.asamblea.go.cr/Centro_de_informacion/"
    "Consultas_SIL/SitePages/SIL.aspx"
)

CARPETA_DOCS        = "documentos"
_mp                 = os.getenv("MAX_PAGINAS", "1")
MAX_PAGINAS         = int(_mp) if _mp.strip() else None   # None = all pages
DESCARGAR_DOCS      = False
TEXTO_BOTON_ENTRADA = "Expedientes Legislativos - Consulta"

ESPERA_CARGA_GRILLA = 4_000
ESPERA_CLIC_FILA    = 2_500
ESPERA_CLIC_TAB     = 1_800
ESPERA_CLIC_PAGINA  = 3_000

# ----------------------------------------------------------------------
# ARGUMENT PARSING (para rangos de páginas / checkpoint)
# ----------------------------------------------------------------------

_parser = argparse.ArgumentParser(add_help=False)
_parser.add_argument("--pagina-inicio", type=int, default=None)
_parser.add_argument("--pagina-fin",    type=int, default=None)
_args, _ = _parser.parse_known_args()

PAGINA_INICIO = _args.pagina_inicio  # None = leer desde checkpoint.json
PAGINA_FIN    = _args.pagina_fin     # None = sin límite de fin

# ----------------------------------------------------------------------
# CHECKPOINT
# ----------------------------------------------------------------------

CHECKPOINT_FILE = "checkpoint.json"
PAGINAS_POR_RUN = 100
LIMITE_MONITOREO = 1200  # Después de la 1era pasada, solo monitoreamos hasta aquí.


def leer_checkpoint():
    """
    Reads the starting page and pass status from checkpoint.json.
    """
    try:
        with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        pagina = int(data.get("pagina_actual", 1))
        completa = data.get("primera_pasada_completa", False)
        print(f"Checkpoint loaded: page {pagina}, first pass complete: {completa}.")
        return pagina, completa
    except Exception:
        print("No checkpoint found. Starting fresh.")
        return 1, False


def guardar_checkpoint(pagina_siguiente: int, ciclo_completado: bool = False, primera_pasada_completa: bool = False):
    """
    Saves the next starting page to checkpoint.json.
    """
    if ciclo_completado:
        pagina_siguiente = 1
        primera_pasada_completa = True  # Al completar cualquier ciclo, asumimos que la 1era ya pasó
        print("Cycle complete. Redirecting to page 1 for next run.")
    
    data = {
        "pagina_actual":           pagina_siguiente,
        "primera_pasada_completa": primera_pasada_completa,
        "ultima_ejecucion":        datetime.utcnow().isoformat(),
    }
    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    print(f"Checkpoint saved: page {pagina_siguiente}.")


# ----------------------------------------------------------------------
# UTILITIES
# ----------------------------------------------------------------------
async def esperar_frame_webpart(page: Page, timeout_ms: int = 30000) -> bool:
    """
    Espera a que el frame MSOPageViewerWebPart tenga contenido cargado.
    """
    inicio = asyncio.get_event_loop().time()
    while (asyncio.get_event_loop().time() - inicio) * 1000 < timeout_ms:
        for frame in page.frames:
            nombre = getattr(frame, 'name', '')
            if 'MSOPageViewer' in nombre or 'WebPartWPQ' in nombre:
                try:
                    count = await frame.evaluate("document.querySelectorAll('a, button').length")
                    if count > 0:
                        print(f"Frame '{nombre}' ready with {count} elements.")
                        return True
                except Exception:
                    pass
        await page.wait_for_timeout(1000)
    return False

def limpiar(v):
    """
    Cleans special characters and control codes from a string.
    """
    if not isinstance(v, str):
        return v
    v = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', v)
    v = re.sub(r'[\u2400-\u243F]', '-', v)
    return v.strip()


# ----------------------------------------------------------------------
# INITIAL NAVIGATION
# ----------------------------------------------------------------------

async def navegar_a_expedientes(page: Page) -> bool:
    titulo_actual = await page.title()
    print(f"Current page title: '{titulo_actual}'")
    
    is_ci = os.getenv("CI", "false").lower() == "true"
    espera_inicial = 8000 if is_ci else 2000
    
    print(f"Waiting {espera_inicial}ms for frames to initialize...")
    await page.wait_for_timeout(espera_inicial)

    print("Waiting for SharePoint WebPart frame to load...")
    frame_listo = await esperar_frame_webpart(page, timeout_ms=45000)
    if not frame_listo:
        print("WebPart frame never loaded content — will try anyway with JS fallback.")
    
    print(f"Searching for button '{TEXTO_BOTON_ENTRADA}'...")
    
    print(f"Searching for button '{TEXTO_BOTON_ENTRADA}' in {len(page.frames)} frames...")
    
    for intento in range(5):
        await page.screenshot(path=f"debug_intento_{intento+1}.png", full_page=True)
        
        frames_a_buscar = [page] + list(page.frames)
        print(f"  Attempt {intento+1}: checking {len(frames_a_buscar)} frames...")
        
        for ctx in frames_a_buscar:
            ctx_name = getattr(ctx, 'name', 'main-page')
            try:
                await ctx.wait_for_load_state("domcontentloaded", timeout=5000)
            except Exception:
                pass
                
            try:
                botones = await ctx.query_selector_all(
                    "a[role='button'], button, a, div[role='button'], span[role='button']"
                )
                print(f"    Frame '{ctx_name}': {len(botones)} clickable elements found")
                
                for boton in botones:
                    try:
                        texto = (await boton.inner_text()).strip()
                        if not texto:
                            texto = await boton.get_attribute("title") or ""
                        if not texto:
                            texto = await boton.get_attribute("aria-label") or ""
                        
                        if TEXTO_BOTON_ENTRADA.lower() in texto.lower():
                            print(f"Button found in frame '{ctx_name}' (attempt {intento+1}). Clicking...")
                            await boton.scroll_into_view_if_needed()
                            await boton.click()
                            await page.wait_for_load_state("networkidle", timeout=45000)
                            await page.wait_for_timeout(ESPERA_CARGA_GRILLA)
                            return True
                    except Exception:
                        continue
            except Exception as e:
                print(f"    Error in frame '{ctx_name}': {e}")
                continue

        # --- NUEVO: fallback ejecutando JS dentro de cada frame ---
        print(f"  Trying per-frame JavaScript fallback on attempt {intento+1}...")
        for ctx in [page] + list(page.frames):
            ctx_name = getattr(ctx, 'name', 'main-page')
            try:
                clicked = await ctx.evaluate(f"""() => {{
                    const texto_buscado = "{TEXTO_BOTON_ENTRADA}".toLowerCase();
                    const selectores = ['a', 'button', 'div[role="button"]', 'span[role="button"]'];
                    for (const sel of selectores) {{
                        for (const el of document.querySelectorAll(sel)) {{
                            const texto = (el.innerText || el.textContent || 
                                           el.getAttribute('title') || 
                                           el.getAttribute('aria-label') || '').toLowerCase().trim();
                            if (texto.includes(texto_buscado)) {{
                                el.click();
                                return true;
                            }}
                        }}
                    }}
                    return false;
                }}""")
                
                if clicked:
                    print(f"Button found via JS in frame '{ctx_name}' (attempt {intento+1}).")
                    await page.wait_for_load_state("networkidle", timeout=45000)
                    await page.wait_for_timeout(ESPERA_CARGA_GRILLA)
                    return True
            except Exception as e:
                print(f"  JS eval error in frame '{ctx_name}': {e}")
                continue

        espera = 8000 + (intento * 4000)
        print(f"Button not found in attempt {intento+1}. Waiting {espera//1000}s before retry...")
        
        if intento == 4:
            with open(f"debug_intento_{intento+1}.html", "w", encoding="utf-8") as f:
                f.write(await page.content())
        
        await page.wait_for_timeout(espera)

    print("Button not found after 5 attempts.")
    return False


async def encontrar_frame_con_grilla(page: Page):
    """
    Finds the frame containing the jqxGrid.
    """
    selectores = ["div[role='grid']", ".jqx-grid", "div.jqx-grid-cell"]
    for ctx in [page] + list(page.frames):
        try:
            for sel in selectores:
                if await ctx.query_selector(sel):
                    print("Data grid located.")
                    return ctx
        except Exception:
            continue
    return None

async def cambiar_mostrar_registros(frame, page: Page, cantidad: str = "10"):
    """
    Locates the 'Show records' dropdown and selects the desired amount.
    """
    print(f"Attempting to change records per page to {cantidad}...")
    try:
        selector_dropdown = ".jqx-dropdownlist-content"
        dropdowns = await frame.query_selector_all(selector_dropdown)
        
        target_dropdown = None
        for d in dropdowns:
            texto = await d.inner_text()
            if texto.strip().isdigit():
                target_dropdown = d
                break
        
        if target_dropdown:
            print(f"Dropdown found (current value: {await target_dropdown.inner_text()}).")
            await target_dropdown.click()
            await page.wait_for_timeout(1000)

            opcion_selector = f".jqx-listitem-element:has-text('{cantidad}')"
            opcion = await frame.query_selector(opcion_selector)
            if not opcion:
                opcion = await page.query_selector(opcion_selector)
            
            if opcion:
                await opcion.click()
                print(f"Selected: {cantidad} records per page.")
                await page.wait_for_timeout(ESPERA_CARGA_GRILLA)
                return True
            else:
                print(f"Option '{cantidad}' not found in dropdown.")
        else:
            print("Records per page control not found.")
            
    except Exception as e:
        print(f"Error changing record count: {e}")
    return False


# ----------------------------------------------------------------------
# READ ROWS
# ----------------------------------------------------------------------

async def obtener_info_filas(frame) -> list:
    """
    Returns a list of dictionaries with {expediente, titulo, row_index}.
    Excludes rows from the bottom panel.
    """
    info = []
    try:
        grilla_principal = await frame.query_selector("div[role='grid']")
        if not grilla_principal:
            return info
        
        rows = await grilla_principal.query_selector_all("div[role='row']")
        for idx, row in enumerate(rows):
            try:
                celdas = await row.query_selector_all("div[role='gridcell']")
                if len(celdas) < 2:
                    continue
                num    = (await celdas[0].inner_text()).strip()
                titulo = (await celdas[1].inner_text()).strip()
                if num and re.match(r'^\d{4,6}$', num.replace(" ", "")):
                    info.append({
                        "expediente": num,
                        "titulo":     titulo,
                        "row_index":  idx,
                    })
            except Exception:
                continue
    except Exception as exc:
        print(f"Error reading grid: {exc}")
    return info


async def clicar_fila_por_indice(frame, page: Page, row_index: int) -> bool:
    """
    Clicks a row in the grid using its index.
    """
    try:
        rows = await frame.query_selector_all("div[role='row']")
        if row_index >= len(rows):
            print(f"Index {row_index} out of range ({len(rows)} rows).")
            return False
        await rows[row_index].click()
        await page.wait_for_timeout(ESPERA_CLIC_FILA)
        return True
    except Exception as exc:
        print(f"Error clicking row {row_index}: {exc}")
        return False


# ----------------------------------------------------------------------
# TAB OPERATIONS
# ----------------------------------------------------------------------

async def clic_tab(frame, page: Page, texto_tab: str) -> bool:
    """
    Clicks a specific tab in the details panel.
    """
    selectores = [
        f"td:has-text('{texto_tab}')",
        f"li[role='tab']:has-text('{texto_tab}')",
        f"div[role='tab']:has-text('{texto_tab}')",
        f"a:has-text('{texto_tab}')",
        f"span:has-text('{texto_tab}')",
    ]
    for sel in selectores:
        try:
            els = await frame.query_selector_all(sel)
            for el in els:
                if await el.is_visible():
                    await el.click()
                    await page.wait_for_timeout(ESPERA_CLIC_TAB)
                    return True
        except Exception:
            continue
    print(f"Tab '{texto_tab}' not found.")
    return False


# ----------------------------------------------------------------------
# EXTRACTION HELPERS
# ----------------------------------------------------------------------

async def extraer_tab_general(frame, page: Page) -> dict:
    """
    Extracts key/value pairs from the 'General' tab in the bottom panel.
    """
    await clic_tab(frame, page, "General")
    await page.wait_for_timeout(500)

    datos = {}
    try:
        datos = await frame.evaluate("""() => {
            const resultado = {};

            for (const tabla of document.querySelectorAll('table')) {
                if (tabla.closest("[role='grid']") || tabla.closest('.jqx-grid')) continue;

                for (const tr of tabla.querySelectorAll('tr')) {
                    const tds = [...tr.querySelectorAll('td')];
                    for (let i = 0; i < tds.length - 1; i++) {
                        const labelTd = tds[i];
                        const valorTd = tds[i + 1];

                        if (labelTd.querySelector('input, select')) continue;

                        const label = (labelTd.innerText || labelTd.textContent || '')
                            .trim().replace(/[:：]$/, '').trim();
                        if (!label || label.length > 60) continue;

                        const inp = valorTd.querySelector('input, select');
                        let val = '';
                        if (inp) {
                            val = (inp.value || inp.getAttribute('value') || '').trim();
                        } else {
                            val = (valorTd.innerText || valorTd.textContent || '').trim();
                        }

                        if (label && val) resultado[label] = val;
                    }
                }
            }

            if (Object.keys(resultado).length === 0) {
                for (const inp of document.querySelectorAll('input[aria-label]')) {
                    const label = (inp.getAttribute('aria-label') || '').trim();
                    const val   = (inp.value || '').trim();
                    if (label && val) resultado[label] = val;
                }
            }

            return resultado;
        }""")
    except Exception as exc:
        print(f"Error in extraer_tab_general: {exc}")

    return {k: limpiar(v) for k, v in datos.items()}

async def _leer_grilla_panel_inferior(frame, page, columnas_esperadas: int,
                                       skip_first: bool = True,
                                       timeout_ms: int = 5000) -> list:
    """
    Internal helper to read ALL rows from the grid inside the details panel,
    scrolling through it to bypass jqxGrid's row virtualization.
    """
    todas_las_filas = []
    filas_vistas = set()

    for scroll_top in range(0, 5000, 150):  # scroll en pasos de 150px hasta 5000px
        resultado = await frame.evaluate(f"""(scrollTop) => {{
            const contenedor = document.querySelector('.marco-subcontenedor.alto-completo');
            if (!contenedor) return null;

            const panelVisible = [...contenedor.querySelectorAll(
                '.jqx-tabs-content-element'
            )].find(p => p.offsetParent !== null);
            if (!panelVisible) return null;

            const grilla = panelVisible.querySelector("div[role='grid']");
            if (!grilla) return null;

            // Hacer scroll dentro del contenedor scrolleable del grid
            const scrollable = grilla.querySelector('.jqx-grid-content') || 
                               grilla.querySelector('[style*="overflow"]') ||
                               grilla;
            scrollable.scrollTop = scrollTop;

            const rows = [...grilla.querySelectorAll("div[role='row']")];
            if (!rows.length) return null;

            const filas = [];
            for (const row of rows) {{
                const celdas = [...row.querySelectorAll("div[role='gridcell']")]
                    .map(c => (c.innerText || c.textContent || '').trim());
                if (celdas.filter(c => c.length > 0).length < 2) continue;
                const slice = {'celdas.slice(1)' if skip_first else 'celdas'};
                filas.push(slice);
            }}
            return filas;
        }}""", scroll_top)

        if resultado is None:
            break

        hubo_nuevas = False
        for fila in resultado:
            clave = tuple(fila)
            if clave not in filas_vistas and any(c for c in fila):
                filas_vistas.add(clave)
                todas_las_filas.append(fila)
                hubo_nuevas = True

        if not hubo_nuevas and scroll_top > 0:
            break  # Ya no aparecen filas nuevas, terminamos

        await page.wait_for_timeout(80)

    if not todas_las_filas:
        print(f"Timeout waiting for bottom grid with {columnas_esperadas} columns.")

    return todas_las_filas

async def extraer_tab_tramitacion(frame, page) -> list:
    await clic_tab(frame, page, "Tramitación")
    await page.wait_for_timeout(800)

    tramitacion = []
    try:
        resultado = await frame.evaluate("""() => {
            const contenedor = document.querySelector('.marco-subcontenedor.alto-completo');
            if (!contenedor) return null;

            const panelVisible = [...contenedor.querySelectorAll(
                '.jqx-tabs-content-element'
            )].find(p => p.offsetParent !== null);
            if (!panelVisible) return null;

            const grilla = panelVisible.querySelector("div[role='grid']");
            if (!grilla) return null;

            const $ = window.$ || window.jQuery;
            if (!$ || !$(grilla).jqxGrid) return null;

            const rowCount = $(grilla).jqxGrid('getdatainformation').rowscount;
            if (rowCount === 0) return [];

            const filas = [];
            for (let i = 0; i < rowCount; i++) {
                const row = $(grilla).jqxGrid('getrowdata', i);
                if (!row) continue;
                filas.push({
                    organo:        String(row.Nombre_Corto       ?? ''),
                    descripcion:   String(row.Descripcion_Tramite ?? ''),
                    fecha_inicio:  String(row.Fecha_Inicio        ?? '').split(' ')[0],
                    fecha_termino: String(row.Fecha_Termino       ?? '').split(' ')[0],
                });
            }
            return filas;
        }""")

        if resultado:
            for fila in resultado:
                tramitacion.append({
                    "Órgano":        limpiar(fila["organo"]),
                    "Descripción":   limpiar(fila["descripcion"]),
                    "Fecha Inicio":  limpiar(fila["fecha_inicio"]),
                    "Fecha Término": limpiar(fila["fecha_termino"]),
                })

    except Exception as e:
        print(f"Error extracting tramitacion: {e}")

    return tramitacion

async def extraer_tab_proponentes(frame, page) -> list:
    await clic_tab(frame, page, "Proponentes")
    await page.wait_for_timeout(800)

    proponentes = []
    try:
        resultado = await frame.evaluate("""() => {
            const contenedor = document.querySelector('.marco-subcontenedor.alto-completo');
            if (!contenedor) return null;

            const panelVisible = [...contenedor.querySelectorAll(
                '.jqx-tabs-content-element'
            )].find(p => p.offsetParent !== null);
            if (!panelVisible) return null;

            const grilla = panelVisible.querySelector("div[role='grid']");
            if (!grilla) return null;

            const $ = window.$ || window.jQuery;
            if (!$ || !$(grilla).jqxGrid) return null;

            const rowCount = $(grilla).jqxGrid('getdatainformation').rowscount;
            if (rowCount === 0) return [];

            const filas = [];
            for (let i = 0; i < rowCount; i++) {
                const row = $(grilla).jqxGrid('getrowdata', i);
                if (!row) continue;
                filas.push({
                    firma:          String(row.Secuencia_Firma ?? ''),
                    nombre:         String(row.Nombre         ?? ''),
                    administracion: String(row.Administracion ?? ''),
                });
            }
            return filas;
        }""")

        if resultado:
            for fila in resultado:
                proponentes.append({
                    "Firma":          limpiar(fila["firma"]),
                    "Nombre":         limpiar(fila["nombre"]),
                    "Administración": limpiar(fila["administracion"]),
                })

    except Exception as e:
        print(f"Error extracting proponentes: {e}")

    return proponentes

async def volver_tab_general(frame, page: Page):
    """
    Resets the details panel to the 'General' tab.
    """
    await clic_tab(frame, page, "General")


# ----------------------------------------------------------------------
# PROCESS PAGE
# ----------------------------------------------------------------------

async def procesar_pagina_grilla(page: Page, frame, num_pagina: int, acumulado: list, expedientes_vistos: set) -> bool:
    """
    Iterates through rows in the current page and extracts details for each file.
    """
    await clic_tab(frame, page, "General")
    await page.wait_for_timeout(500)

    filas_info = await obtener_info_filas(frame)

    if not filas_info:
        print(f"No rows found on page {num_pagina}.")
        return False
        
    primer_exp = filas_info[0]["expediente"]
    if primer_exp in expedientes_vistos:
        print(f"Cycle detected: File {primer_exp} already processed. Synchronization complete.")
        return False
        
    expedientes_vistos.add(primer_exp)

    total = len(filas_info)
    print("-" * 60)
    print(f"PAGE {num_pagina} - {total} files found")
    print("-" * 60)

    for i, info in enumerate(filas_info):
        num_exp      = info["expediente"]
        titulo       = info["titulo"]
        row_index    = info["row_index"]
        titulo_corto = limpiar(titulo)[:65] + ("..." if len(titulo) > 65 else "")

        print(f"[{i+1}/{total}] File {num_exp}: {titulo_corto}")

        ok = await clicar_fila_por_indice(frame, page, row_index)
        if not ok:
            continue

        general     = await extraer_tab_general(frame, page)
        tramitacion = await extraer_tab_tramitacion(frame, page)
        proponentes = await extraer_tab_proponentes(frame, page)

        await volver_tab_general(frame, page)
        await page.wait_for_timeout(300)

        acumulado.append({
            "pagina":            num_pagina,
            "numero_expediente": num_exp,
            "titulo":            limpiar(titulo),
            "general":           general,
            "tramitacion":       tramitacion,
            "proponentes":       proponentes,
        })

        print(
            f"      Data extracted: General ({len(general)}), "
            f"Tramitación ({len(tramitacion)}), Proponentes ({len(proponentes)})"
        )
            
    return True


# ----------------------------------------------------------------------
# PAGINATION
# ----------------------------------------------------------------------
async def ir_a_pagina_directa(frame, page: Page, numero_pagina: int) -> bool:
    """
    Navega directamente a una página escribiendo el número en el input 'ctrl-tabla-ira'.
    """
    try:
        # Buscar el input de página actual
        input_pagina = await frame.query_selector("input.ctrl-tabla-ira")
        if not input_pagina:
            input_pagina = await frame.query_selector("input[title='Página actual']")
        
        if not input_pagina:
            print(f"Input de página no encontrado. Usando navegación secuencial.")
            return False

        # Leer el número de página actual antes de cambiar
        filas_antes = await obtener_info_filas(frame)
        primer_exp_antes = filas_antes[0]["expediente"] if filas_antes else None

        # Limpiar, escribir el número y presionar Enter
        await input_pagina.click(click_count=3)
        await input_pagina.fill(str(numero_pagina))
        await input_pagina.press("Enter")
        
        print(f"Navegando directamente a página {numero_pagina}...")
        await page.wait_for_timeout(ESPERA_CLIC_PAGINA)

        # Verificar que la página cambió
        for intento in range(10):
            filas_despues = await obtener_info_filas(frame)
            primer_exp_despues = filas_despues[0]["expediente"] if filas_despues else None

            if primer_exp_despues and primer_exp_despues != primer_exp_antes:
                print(f"Página {numero_pagina} cargada correctamente.")
                return True

            print(f"Esperando cambio de página... intento {intento + 1}/10")
            await page.wait_for_timeout(800)

        print(f"La grilla no cambió al ir a página {numero_pagina}.")
        return False

    except Exception as e:
        print(f"Error navegando a página {numero_pagina}: {e}")
        return False
async def ir_siguiente_pagina(frame, page: Page, pagina_actual: int) -> bool:
    """
    Clicks the next page button and verifies the content has changed.
    """
    filas_antes = await obtener_info_filas(frame)
    primer_exp_antes = filas_antes[0]["expediente"] if filas_antes else None

    selectores = [
        "div[title='Siguiente página']",
        "input[title='Siguiente página']",
        "div[title='Next Page']",
        "input[title='Next Page']",
        ".jqx-icon-arrow-right:not(.jqx-icon-arrow-right-selected)",
    ]
    clic_exitoso = False
    for sel in selectores:
        try:
            btns = await frame.query_selector_all(sel)
            for btn in btns:
                if not await btn.is_visible():
                    continue
                clase    = (await btn.get_attribute("class") or "").lower()
                disabled = await btn.get_attribute("disabled")
                title    = (await btn.get_attribute("title") or "").lower()
                if disabled or "disabled" in clase:
                    continue
                if "ltima" in title or "last" in title:
                    continue
                await btn.click()
                clic_exitoso = True
                break
        except Exception:
            continue
        if clic_exitoso:
            break

    if not clic_exitoso:
        try:
            for btn in await frame.query_selector_all("div, button, a, input[type='button']"):
                try:
                    texto = (await btn.inner_text()).strip()
                    title = (await btn.get_attribute("title") or "").lower()
                    clase = (await btn.get_attribute("class") or "").lower()
                    if (texto in [">", "»", "›"] or "siguiente" in title or "next" in title):
                        if "disabled" not in clase and not await btn.get_attribute("disabled"):
                            await btn.click()
                            clic_exitoso = True
                            break
                except Exception:
                    continue
        except Exception:
            pass

    if not clic_exitoso:
        print("Next page button not found. Assuming end of list.")
        return False

    await page.wait_for_timeout(ESPERA_CLIC_PAGINA)

    for intento in range(10):
        filas_despues = await obtener_info_filas(frame)
        primer_exp_despues = filas_despues[0]["expediente"] if filas_despues else None

        if primer_exp_despues and primer_exp_despues != primer_exp_antes:
            print(f"Navigated to page {pagina_actual + 1}.")
            return True

        print(f"Waiting for page change... attempt {intento + 1}/10")
        await page.wait_for_timeout(800)

    print("Data grid did not change. Possible end of list.")
    return False


# ----------------------------------------------------------------------
# EXPORT
# ----------------------------------------------------------------------

def exportar_excel(proyectos: list, nombre: str):
    """
    Exports captured project data to an Excel workbook.
    """
    h_fill = PatternFill("solid", fgColor="1F4E79")
    h_font = Font(bold=True, color="FFFFFF")
    centro = Alignment(horizontal="center")

    def enc(ws, cols):
        for c, txt in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=txt)
            cell.font = h_font; cell.fill = h_fill; cell.alignment = centro

    def s(v):
        return limpiar(v) if isinstance(v, str) else v

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Resumen"
    enc(ws, ["Página", "Expediente", "Título", "Tipo expediente",
             "Fecha inicio", "Vencimiento cuatrienal", "Ley", "Acuerdo", "Veto"])
    for r, p in enumerate(proyectos, 2):
        g = p.get("general", {})
        ws.cell(r, 1, p.get("pagina"))
        ws.cell(r, 2, s(p.get("numero_expediente", "")))
        ws.cell(r, 3, s(p.get("titulo", "")))
        ws.cell(r, 4, s(g.get("Tipo expediente", "")))
        ws.cell(r, 5, s(g.get("Fecha inicio", "")))
        ws.cell(r, 6, s(g.get("Vencimiento cuatrienal", "")))
        ws.cell(r, 7, s(g.get("Ley", "")))
        ws.cell(r, 8, s(g.get("Acuerdo", "")))
        ws.cell(r, 9, s(g.get("Veto", "")))
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 40

    ws_g = wb.create_sheet("General")
    enc(ws_g, ["Expediente", "Campo", "Valor"])
    r = 2
    for p in proyectos:
        for campo, val in p.get("general", {}).items():
            ws_g.cell(r, 1, s(p.get("numero_expediente", "")))
            ws_g.cell(r, 2, s(campo))
            ws_g.cell(r, 3, s(val))
            r += 1
    ws_g.column_dimensions["B"].width = 35
    ws_g.column_dimensions["C"].width = 55

    ws_t = wb.create_sheet("Tramitación")
    enc(ws_t, ["Expediente", "Órgano", "Descripción", "Fecha Inicio", "Fecha Término"])
    r = 2
    for p in proyectos:
        for t in p.get("tramitacion", []):
            ws_t.cell(r, 1, s(p.get("numero_expediente", "")))
            ws_t.cell(r, 2, s(t.get("Órgano", "")))
            ws_t.cell(r, 3, s(t.get("Descripción", "")))
            ws_t.cell(r, 4, s(t.get("Fecha Inicio", "")))
            ws_t.cell(r, 5, s(t.get("Fecha Término", "")))
            r += 1
    ws_t.column_dimensions["C"].width = 50

    ws_p = wb.create_sheet("Proponentes")
    enc(ws_p, ["Expediente", "Firma", "Nombre", "Administración"])
    r = 2
    for p in proyectos:
        for prop in p.get("proponentes", []):
            ws_p.cell(r, 1, s(p.get("numero_expediente", "")))
            ws_p.cell(r, 2, s(prop.get("Firma", "")))
            ws_p.cell(r, 3, s(prop.get("Nombre", "")))
            ws_p.cell(r, 4, s(prop.get("Administración", "")))
            r += 1
    ws_p.column_dimensions["C"].width = 40

    wb.save(nombre)
    print(f"Excel report saved: {nombre}")


# ----------------------------------------------------------------------
# MAIN
# ----------------------------------------------------------------------

async def main():
    # Determinar página de inicio: argumento CLI tiene prioridad, luego checkpoint
    if PAGINA_INICIO is not None:
        pagina_inicio = PAGINA_INICIO
        primera_pasada_completa = False
    else:
        pagina_inicio, primera_pasada_completa = leer_checkpoint()
    
    pagina_fin = PAGINA_FIN if PAGINA_FIN is not None else pagina_inicio + PAGINAS_POR_RUN - 1

    inicio = datetime.now()

    print("-" * 60)
    print("SIL Extractor - Legislative Assembly of Costa Rica")
    print("-" * 60)
    print(f"Started at:    {inicio:%Y-%m-%d %H:%M:%S}")
    print(f"Page range:    {pagina_inicio} → {pagina_fin}")
    print(f"Scope:         {'ALL PAGES' if not MAX_PAGINAS else f'first {MAX_PAGINAS} pages'}")
    print("-" * 60)

    os.makedirs(CARPETA_DOCS, exist_ok=True)
    proyectos = []

    async with async_playwright() as p:
        is_ci = os.getenv("CI", "false").lower() == "true"
        browser = await p.chromium.launch(
            headless=is_ci,
            args=["--no-sandbox", "--disable-dev-shm-usage",
                  "--disable-blink-features=AutomationControlled"]
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1440, "height": 900},
        )
        page = await context.new_page()

        print("Loading SIL portal...")
        try:
            await page.goto(URL_BASE, wait_until="networkidle", timeout=60_000)
        except Exception as e:
            print(f"Warning: Initial load timeout or networkidle issue: {e}")
            await page.goto(URL_BASE, wait_until="domcontentloaded", timeout=30_000)
        
        await page.wait_for_timeout(8_000)

        if not await navegar_a_expedientes(page):
            print("Failed to navigate to the module.")
            await browser.close()
            return

        print("Locating data grid...")
        await page.wait_for_timeout(3_000)
        frame = await encontrar_frame_con_grilla(page)

        if not frame:
            print("Grid not found. Aborting.")
            await browser.close()
            return

        await cambiar_mostrar_registros(frame, page, "10")
        await page.wait_for_timeout(6_000)

        # Avanzar hasta la página de inicio si no es la 1
        if pagina_inicio > 1:
            print(f"Saltando directamente a página {pagina_inicio}...")
            if not await ir_a_pagina_directa(frame, page, pagina_inicio):
                print(f"No se pudo llegar a la página {pagina_inicio}. Abortando.")
                await browser.close()
                return
            print(f"Listo. Iniciando desde página {pagina_inicio}.")

        pagina_actual      = pagina_inicio
        paginas_procesadas = 0
        expedientes_vistos = set()
        ciclo_finalizado   = False

        while True:
            # 1. Límite de fase de monitoreo (si ya se hizo la 1era pasada)
            if primera_pasada_completa and pagina_actual > LIMITE_MONITOREO:
                print(f"Monitoring limit reached ({LIMITE_MONITOREO}). Resetting cycle.")
                ciclo_finalizado = True
                break

            if MAX_PAGINAS and paginas_procesadas >= MAX_PAGINAS:
                print(f"Limit of {MAX_PAGINAS} pages reached.")
                break

            if pagina_actual > pagina_fin:
                print(f"Batch complete: pages {pagina_inicio}-{pagina_fin}.")
                break

            continuar = await procesar_pagina_grilla(page, frame, pagina_actual, proyectos, expedientes_vistos)
            if not continuar:
                ciclo_finalizado = True
                break

            paginas_procesadas += 1

            print(f"Attempting to advance to page {pagina_actual + 1}...")
            if not await ir_siguiente_pagina(frame, page, pagina_actual):
                print("No more pages available. Full sync complete.")
                ciclo_finalizado = True
                break

            pagina_actual += 1

        await browser.close()

    # Guardar checkpoint para el próximo run
    guardar_checkpoint(
        pagina_actual, 
        ciclo_completado=ciclo_finalizado, 
        primera_pasada_completa=primera_pasada_completa
    )

    if not proyectos:
        print("No projects were extracted.")
        return

    print("-" * 60)
    print("DATABASE SYNCHRONIZATION")
    print("-" * 60)
    crear_tablas()
    stats = sync_proyectos(proyectos)

    ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
    rango      = f"p{pagina_inicio}-{pagina_fin}"
    json_file  = f"proyectos_{ts}_{rango}.json"
    excel_file = f"proyectos_{ts}_{rango}.xlsx"

    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(proyectos, f, ensure_ascii=False, indent=2)
    print(f"JSON data saved: {json_file}")

    exportar_excel(proyectos, excel_file)

    duracion = datetime.now() - inicio

    print("-" * 60)
    print("FINAL SUMMARY")
    print("-" * 60)
    print(f"Projects extracted:   {len(proyectos)}")
    print(f"Pages processed:      {paginas_procesadas}")
    print(f"Page range:           {pagina_inicio}-{pagina_fin}")
    print(f"Database - Sync:      {stats.get('actualizados', 0)}")
    print(f"Database - Errors:    {stats.get('errores', 0)}")
    print(f"JSON Output:          {json_file}")
    print(f"Excel Output:         {excel_file}")
    print(f"Total Duration:       {str(duracion).split('.')[0]}")
    print("-" * 60)


if __name__ == "__main__":
    asyncio.run(main())