"""
fase1_scraper.py
──────────────────────────────────────────────────────────────────────
Fase 1 — Datos urgentes del día.

Responsabilidad única: extraer las primeras 3 páginas del SIL de la
Asamblea Legislativa de Costa Rica y sincronizarlas contra PostgreSQL.

Diseñado para correr en GitHub Actions todos los días (~10-15 min).
Si el portal falla en cualquier punto → se detiene limpiamente y sale
con código 0 (no falla el workflow, solo no hay datos ese día).

Flujo:
  1. Navegar al portal y encontrar la grilla
  2. Para cada página 1-3:
     a. Leer filas
     b. Para cada fila → extraer General, Tramitación, Proponentes
  3. Sync contra PostgreSQL via sync_engine.py
  4. Guardar reporte JSON + Excel de lo extraído
"""

import asyncio
import json
import os
import re
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from playwright.async_api import async_playwright, Page

from sync_engine import crear_tablas, sync_proyectos

# ──────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ──────────────────────────────────────────────────────────────────────

URL_BASE = (
    "https://www.asamblea.go.cr/Centro_de_informacion/"
    "Consultas_SIL/SitePages/SIL.aspx"
)

TEXTO_BOTON_ENTRADA = "Expedientes Legislativos - Consulta"
PAGINAS_A_EXTRAER   = 10      # Siempre páginas 1, 2 y 3
REGISTROS_POR_PAG   = "10"   # Valor del dropdown de la grilla

# Tiempos de espera (ms) — ajustados para ser conservadores
ESPERA_CARGA_GRILLA = 5_000
ESPERA_CLIC_FILA    = 2_500
ESPERA_CLIC_TAB     = 1_500
ESPERA_CLIC_PAGINA  = 4_000

IS_CI = os.getenv("CI", "false").lower() == "true"


# ──────────────────────────────────────────────────────────────────────
# UTILIDADES
# ──────────────────────────────────────────────────────────────────────

def limpiar(v):
    """Elimina caracteres de control de una cadena."""
    if not isinstance(v, str):
        return v
    v = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', v)
    v = re.sub(r'[\u2400-\u243F]', '-', v)
    return v.strip()


def log(msg: str):
    """Print con timestamp."""
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


# ──────────────────────────────────────────────────────────────────────
# NAVEGACIÓN INICIAL
# ──────────────────────────────────────────────────────────────────────

async def esperar_frame_webpart(page: Page, timeout_ms: int = 60_000) -> bool:
    """Espera a que el frame del WebPart de SharePoint tenga contenido."""
    inicio = asyncio.get_event_loop().time()
    while (asyncio.get_event_loop().time() - inicio) * 1000 < timeout_ms:
        for frame in page.frames:
            nombre = getattr(frame, 'name', '')
            if 'MSOPageViewer' in nombre or 'WebPartWPQ' in nombre:
                try:
                    count = await frame.evaluate(
                        "document.querySelectorAll('a, button').length"
                    )
                    if count > 0:
                        log(f"Frame '{nombre}' listo ({count} elementos).")
                        return True
                except Exception:
                    pass
        await page.wait_for_timeout(1_000)
    return False


async def navegar_a_expedientes(page: Page) -> bool:
    """
    Intenta navegar al módulo de expedientes legislativos.
    Máximo 5 intentos con 60s de espera entre cada uno.
    Retorna True si logra hacer clic en el botón de entrada.
    """
    for intento in range(5):
        if intento > 0:
            log(f"Reintento {intento}/4 — recargando página...")
            try:
                await page.goto(URL_BASE, wait_until="networkidle", timeout=60_000)
            except Exception:
                await page.goto(URL_BASE, wait_until="domcontentloaded", timeout=30_000)

        log("Esperando frame del portal...")
        if not await esperar_frame_webpart(page, timeout_ms=60_000):
            log(f"Frame no cargó en 60s. Esperando 60s antes de reintentar...")
            await page.screenshot(path=f"debug_fase1_intento_{intento+1}.png", full_page=True)
            await page.wait_for_timeout(60_000)
            continue

        log(f"Buscando botón '{TEXTO_BOTON_ENTRADA}'...")
        contextos = [page] + list(page.frames)

        # Intento directo con Playwright
        for ctx in contextos:
            ctx_name = getattr(ctx, 'name', 'main')
            try:
                selectores = [
                    "a[role='button']", "button", "a",
                    "div[role='button']", "span[role='button']"
                ]
                for sel in selectores:
                    botones = await ctx.query_selector_all(sel)
                    for boton in botones:
                        try:
                            texto = (await boton.inner_text()).strip()
                            if not texto:
                                texto = await boton.get_attribute("title") or ""
                            if not texto:
                                texto = await boton.get_attribute("aria-label") or ""
                            if TEXTO_BOTON_ENTRADA.lower() in texto.lower():
                                log(f"Botón encontrado en frame '{ctx_name}'. Haciendo clic...")
                                await boton.scroll_into_view_if_needed()
                                await boton.click()
                                await page.wait_for_load_state("networkidle", timeout=45_000)
                                await page.wait_for_timeout(ESPERA_CARGA_GRILLA)
                                return True
                        except Exception:
                            continue
            except Exception:
                continue

        # Fallback JS por frame
        for ctx in contextos:
            ctx_name = getattr(ctx, 'name', 'main')
            try:
                clicked = await ctx.evaluate(f"""() => {{
                    const texto_buscado = "{TEXTO_BOTON_ENTRADA}".toLowerCase();
                    for (const sel of ['a','button','div[role="button"]','span[role="button"]']) {{
                        for (const el of document.querySelectorAll(sel)) {{
                            const t = (el.innerText || el.textContent ||
                                       el.getAttribute('title') ||
                                       el.getAttribute('aria-label') || '').toLowerCase().trim();
                            if (t.includes(texto_buscado)) {{ el.click(); return true; }}
                        }}
                    }}
                    return false;
                }}""")
                if clicked:
                    log(f"Botón encontrado via JS en frame '{ctx_name}'.")
                    await page.wait_for_load_state("networkidle", timeout=45_000)
                    await page.wait_for_timeout(ESPERA_CARGA_GRILLA)
                    return True
            except Exception:
                continue

        log("Botón no encontrado. Guardando debug...")
        await page.screenshot(path=f"debug_fase1_intento_{intento+1}.png", full_page=True)
        with open(f"debug_fase1_intento_{intento+1}.html", "w", encoding="utf-8") as f:
            f.write(await page.content())
        await page.wait_for_timeout(60_000)

    log("No se pudo navegar al módulo tras 5 intentos.")
    return False


async def encontrar_frame_con_grilla(page: Page):
    """Encuentra el frame que contiene el jqxGrid."""
    selectores = ["div[role='grid']", ".jqx-grid", "div.jqx-grid-cell"]
    for ctx in [page] + list(page.frames):
        try:
            for sel in selectores:
                if await ctx.query_selector(sel):
                    log("Grilla localizada.")
                    return ctx
        except Exception:
            continue
    return None


async def cambiar_registros_por_pagina(frame, page: Page, cantidad: str = "10") -> bool:
    """Selecciona la cantidad de registros por página en el dropdown."""
    log(f"Configurando {cantidad} registros por página...")
    try:
        dropdowns = await frame.query_selector_all(".jqx-dropdownlist-content")
        for d in dropdowns:
            texto = await d.inner_text()
            if texto.strip().isdigit():
                await d.click()
                await page.wait_for_timeout(800)
                for ctx in [frame, page]:
                    opcion = await ctx.query_selector(
                        f".jqx-listitem-element:has-text('{cantidad}')"
                    )
                    if opcion:
                        await opcion.click()
                        log(f"Dropdown: {cantidad} registros seleccionados.")
                        await page.wait_for_timeout(ESPERA_CARGA_GRILLA)
                        return True
    except Exception as e:
        log(f"Error configurando registros por página: {e}")
    return False


# ──────────────────────────────────────────────────────────────────────
# MODALES DE ERROR
# ──────────────────────────────────────────────────────────────────────

async def cerrar_modal_error(frame, page: Page) -> bool:
    """
    Detecta y cierra diálogos de error jqx-window-modal.
    Retorna True si había un modal y fue cerrado.
    """
    try:
        modal = await frame.query_selector(".jqx-window-modal")
        if not modal or not await modal.is_visible():
            return False
        log("Modal de error detectado. Cerrando...")
        for sel in [
            ".jqx-window-close-button",
            "button:has-text('Cerrar')",
            "button:has-text('Aceptar')",
            "button:has-text('OK')",
        ]:
            try:
                btn = await frame.query_selector(sel)
                if btn and await btn.is_visible():
                    await btn.click()
                    await page.wait_for_timeout(600)
                    return True
            except Exception:
                continue
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(600)
        return True
    except Exception:
        return False


# ──────────────────────────────────────────────────────────────────────
# LECTURA DE FILAS
# ──────────────────────────────────────────────────────────────────────

async def obtener_info_filas(frame) -> list:
    """
    Lee las filas de la grilla principal.
    Retorna lista de {expediente, titulo, row_index}.
    """
    info = []
    try:
        grilla = await frame.query_selector("div[role='grid']")
        if not grilla:
            return info
        rows = await grilla.query_selector_all("div[role='row']")
        for idx, row in enumerate(rows):
            try:
                celdas = await row.query_selector_all("div[role='gridcell']")
                if len(celdas) < 2:
                    continue
                num    = (await celdas[0].inner_text()).strip()
                titulo = (await celdas[1].inner_text()).strip()
                if num and re.match(r'^\d{4,6}$', num.replace(" ", "")):
                    info.append({
                        "expediente": num,
                        "titulo":     titulo,
                        "row_index":  idx,
                    })
            except Exception:
                continue
    except Exception as exc:
        log(f"Error leyendo grilla: {exc}")
    return info


async def clicar_fila(frame, page: Page, row_index: int) -> bool:
    """Hace clic en una fila. Maneja modales antes y después del clic."""
    await cerrar_modal_error(frame, page)
    try:
        rows = await frame.query_selector_all("div[role='row']")
        if row_index >= len(rows):
            return False
        await rows[row_index].click()
        await page.wait_for_timeout(ESPERA_CLIC_FILA)
        if await cerrar_modal_error(frame, page):
            log(f"  Fila {row_index}: modal de error del servidor. Saltando.")
            return False
        return True
    except Exception as exc:
        log(f"  Error haciendo clic en fila {row_index}: {exc}")
        await cerrar_modal_error(frame, page)
        return False


# ──────────────────────────────────────────────────────────────────────
# TABS DE DETALLE
# ──────────────────────────────────────────────────────────────────────

async def clic_tab(frame, page: Page, texto: str) -> bool:
    """Hace clic en un tab del panel de detalle."""
    for sel in [
        f"td:has-text('{texto}')",
        f"li[role='tab']:has-text('{texto}')",
        f"div[role='tab']:has-text('{texto}')",
        f"a:has-text('{texto}')",
        f"span:has-text('{texto}')",
    ]:
        try:
            els = await frame.query_selector_all(sel)
            for el in els:
                if await el.is_visible():
                    await el.click()
                    await page.wait_for_timeout(ESPERA_CLIC_TAB)
                    return True
        except Exception:
            continue
    return False


async def extraer_tab_general(frame, page: Page) -> dict:
    """Extrae los campos clave/valor del tab General."""
    await clic_tab(frame, page, "General")
    await page.wait_for_timeout(400)
    datos = {}
    try:
        datos = await frame.evaluate("""() => {
            const resultado = {};
            for (const tabla of document.querySelectorAll('table')) {
                if (tabla.closest("[role='grid']") || tabla.closest('.jqx-grid')) continue;
                for (const tr of tabla.querySelectorAll('tr')) {
                    const tds = [...tr.querySelectorAll('td')];
                    for (let i = 0; i < tds.length - 1; i++) {
                        const labelTd = tds[i];
                        const valorTd = tds[i + 1];
                        if (labelTd.querySelector('input, select')) continue;
                        const label = (labelTd.innerText || labelTd.textContent || '')
                            .trim().replace(/[:：]$/, '').trim();
                        if (!label || label.length > 60) continue;
                        const inp = valorTd.querySelector('input, select');
                        let val = '';
                        if (inp) {
                            val = (inp.value || inp.getAttribute('value') || '').trim();
                        } else {
                            val = (valorTd.innerText || valorTd.textContent || '').trim();
                        }
                        if (label && val) resultado[label] = val;
                    }
                }
            }
            if (Object.keys(resultado).length === 0) {
                for (const inp of document.querySelectorAll('input[aria-label]')) {
                    const label = (inp.getAttribute('aria-label') || '').trim();
                    const val   = (inp.value || '').trim();
                    if (label && val) resultado[label] = val;
                }
            }
            return resultado;
        }""")
    except Exception as exc:
        log(f"  Error en tab General: {exc}")
    return {k: limpiar(v) for k, v in datos.items()}


async def extraer_tab_tramitacion(frame, page: Page) -> list:
    """Extrae los pasos de tramitación via jqxGrid API."""
    await clic_tab(frame, page, "Tramitación")
    await page.wait_for_timeout(700)
    tramitacion = []
    try:
        resultado = await frame.evaluate("""() => {
            const contenedor = document.querySelector('.marco-subcontenedor.alto-completo');
            if (!contenedor) return null;
            const panelVisible = [...contenedor.querySelectorAll('.jqx-tabs-content-element')]
                .find(p => p.offsetParent !== null);
            if (!panelVisible) return null;
            const grilla = panelVisible.querySelector("div[role='grid']");
            if (!grilla) return null;
            const $ = window.$ || window.jQuery;
            if (!$ || !$(grilla).jqxGrid) return null;
            const rowCount = $(grilla).jqxGrid('getdatainformation').rowscount;
            if (rowCount === 0) return [];
            const filas = [];
            for (let i = 0; i < rowCount; i++) {
                const row = $(grilla).jqxGrid('getrowdata', i);
                if (!row) continue;
                filas.push({
                    organo:        String(row.Nombre_Corto        ?? ''),
                    descripcion:   String(row.Descripcion_Tramite ?? ''),
                    fecha_inicio:  String(row.Fecha_Inicio        ?? '').split(' ')[0],
                    fecha_termino: String(row.Fecha_Termino       ?? '').split(' ')[0],
                });
            }
            return filas;
        }""")
        if resultado:
            for f in resultado:
                tramitacion.append({
                    "Órgano":        limpiar(f["organo"]),
                    "Descripción":   limpiar(f["descripcion"]),
                    "Fecha Inicio":  limpiar(f["fecha_inicio"]),
                    "Fecha Término": limpiar(f["fecha_termino"]),
                })
    except Exception as e:
        log(f"  Error en tab Tramitación: {e}")
    return tramitacion


async def extraer_tab_proponentes(frame, page: Page) -> list:
    """Extrae los proponentes via jqxGrid API."""
    await clic_tab(frame, page, "Proponentes")
    await page.wait_for_timeout(700)
    proponentes = []
    try:
        resultado = await frame.evaluate("""() => {
            const contenedor = document.querySelector('.marco-subcontenedor.alto-completo');
            if (!contenedor) return null;
            const panelVisible = [...contenedor.querySelectorAll('.jqx-tabs-content-element')]
                .find(p => p.offsetParent !== null);
            if (!panelVisible) return null;
            const grilla = panelVisible.querySelector("div[role='grid']");
            if (!grilla) return null;
            const $ = window.$ || window.jQuery;
            if (!$ || !$(grilla).jqxGrid) return null;
            const rowCount = $(grilla).jqxGrid('getdatainformation').rowscount;
            if (rowCount === 0) return [];
            const filas = [];
            for (let i = 0; i < rowCount; i++) {
                const row = $(grilla).jqxGrid('getrowdata', i);
                if (!row) continue;
                filas.push({
                    firma:          String(row.Secuencia_Firma ?? ''),
                    nombre:         String(row.Nombre         ?? ''),
                    administracion: String(row.Administracion ?? ''),
                });
            }
            return filas;
        }""")
        if resultado:
            for f in resultado:
                proponentes.append({
                    "Firma":          limpiar(f["firma"]),
                    "Nombre":         limpiar(f["nombre"]),
                    "Administración": limpiar(f["administracion"]),
                })
    except Exception as e:
        log(f"  Error en tab Proponentes: {e}")
    return proponentes


# ──────────────────────────────────────────────────────────────────────
# PAGINACIÓN
# ──────────────────────────────────────────────────────────────────────

async def ir_a_pagina_1(frame, page: Page) -> bool:
    """Navega directamente a la página 1 usando el input de paginación."""
    try:
        inp = await frame.query_selector("input.ctrl-tabla-ira") or \
              await frame.query_selector("input[title='Página actual']")
        if not inp:
            return False
        await inp.click(click_count=3)
        await inp.fill("1")
        await inp.press("Enter")
        await page.wait_for_timeout(ESPERA_CLIC_PAGINA)
        return True
    except Exception as e:
        log(f"Error navegando a página 1: {e}")
        return False


async def ir_siguiente_pagina(frame, page: Page, pagina_actual: int) -> bool:
    """
    Hace clic en 'Siguiente página' y verifica que el contenido cambie.
    Retorna True si la navegación fue exitosa.
    """
    filas_antes = await obtener_info_filas(frame)
    exp_antes = filas_antes[0]["expediente"] if filas_antes else None

    selectores_siguiente = [
        "div[title='Siguiente página']",
        "input[title='Siguiente página']",
        "div[title='Next Page']",
        "input[title='Next Page']",
        ".jqx-icon-arrow-right:not(.jqx-icon-arrow-right-selected)",
    ]
    clic_ok = False
    for sel in selectores_siguiente:
        try:
            for btn in await frame.query_selector_all(sel):
                if not await btn.is_visible():
                    continue
                clase    = (await btn.get_attribute("class") or "").lower()
                disabled = await btn.get_attribute("disabled")
                title    = (await btn.get_attribute("title") or "").lower()
                if disabled or "disabled" in clase:
                    continue
                if "ltima" in title or "last" in title:
                    continue
                await btn.click()
                clic_ok = True
                break
        except Exception:
            continue
        if clic_ok:
            break

    # ── Fallback: buscar por texto o título en todos los elementos ──
    if not clic_ok:
        try:
            for btn in await frame.query_selector_all("div, button, a, input[type='button']"):
                try:
                    texto = (await btn.inner_text()).strip()
                    title = (await btn.get_attribute("title") or "").lower()
                    clase = (await btn.get_attribute("class") or "").lower()
                    if (texto in [">", "»", "›"] or "siguiente" in title or "next" in title):
                        if "disabled" not in clase and not await btn.get_attribute("disabled"):
                            await btn.click()
                            clic_ok = True
                            break
                except Exception:
                    continue
        except Exception:
            pass

    if not clic_ok:
        log(f"Botón 'Siguiente página' no encontrado o deshabilitado.")
        return False

    await page.wait_for_timeout(ESPERA_CLIC_PAGINA)

    for _ in range(10):
        filas_despues = await obtener_info_filas(frame)
        exp_despues = filas_despues[0]["expediente"] if filas_despues else None
        if exp_despues and exp_despues != exp_antes:
            log(f"Navegado a página {pagina_actual + 1} correctamente.")
            return True
        await page.wait_for_timeout(700)

    log("La grilla no cambió al hacer clic en siguiente.")
    return False


# ──────────────────────────────────────────────────────────────────────
# PROCESAR UNA PÁGINA COMPLETA
# ──────────────────────────────────────────────────────────────────────

async def procesar_pagina(
    page: Page,
    frame,
    num_pagina: int,
    acumulado: list,
) -> bool:
    """
    Procesa todas las filas de una página.
    Retorna True si fue exitoso, False si hubo un error que impide continuar.
    """
    # Asegurarse de estar en el tab General antes de leer filas
    await clic_tab(frame, page, "General")
    await page.wait_for_timeout(400)

    filas = await obtener_info_filas(frame)
    if not filas:
        log(f"Página {num_pagina}: sin filas. El portal puede estar con problemas.")
        return False

    total = len(filas)
    log(f"{'─'*50}")
    log(f"PÁGINA {num_pagina} — {total} expedientes")
    log(f"{'─'*50}")

    for i, info in enumerate(filas):
        exp       = info["expediente"]
        titulo    = info["titulo"]
        row_idx   = info["row_index"]
        titulo_c  = limpiar(titulo)[:60] + ("…" if len(titulo) > 60 else "")

        log(f"[{i+1}/{total}] Exp. {exp}: {titulo_c}")

        ok = await clicar_fila(frame, page, row_idx)
        if not ok:
            log(f"  Fila {i+1} omitida (error de clic).")
            continue

        general     = await extraer_tab_general(frame, page)
        tramitacion = await extraer_tab_tramitacion(frame, page)
        proponentes = await extraer_tab_proponentes(frame, page)

        # Volver a General para dejar el panel limpio para la siguiente fila
        await clic_tab(frame, page, "General")
        await page.wait_for_timeout(200)

        acumulado.append({
            "pagina":            num_pagina,
            "numero_expediente": exp,
            "titulo":            limpiar(titulo),
            "general":           general,
            "tramitacion":       tramitacion,
            "proponentes":       proponentes,
        })

        log(
            f"  ✓ General({len(general)}) "
            f"Tramitación({len(tramitacion)}) "
            f"Proponentes({len(proponentes)})"
        )

    return True


# ──────────────────────────────────────────────────────────────────────
# EXPORTAR EXCEL
# ──────────────────────────────────────────────────────────────────────

def exportar_excel(proyectos: list, nombre: str):
    """Exporta los proyectos extraídos a un Excel con 4 hojas."""
    h_fill = PatternFill("solid", fgColor="1F4E79")
    h_font = Font(bold=True, color="FFFFFF")
    centro = Alignment(horizontal="center")

    def enc(ws, cols):
        for c, txt in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=txt)
            cell.font = h_font
            cell.fill = h_fill
            cell.alignment = centro

    def s(v):
        return limpiar(v) if isinstance(v, str) else v

    wb = openpyxl.Workbook()

    # Hoja Resumen
    ws = wb.active
    ws.title = "Resumen"
    enc(ws, ["Página", "Expediente", "Título", "Tipo expediente",
             "Fecha inicio", "Vencimiento cuatrienal", "Ley"])
    for r, p in enumerate(proyectos, 2):
        g = p.get("general", {})
        ws.cell(r, 1, p.get("pagina"))
        ws.cell(r, 2, s(p.get("numero_expediente", "")))
        ws.cell(r, 3, s(p.get("titulo", "")))
        ws.cell(r, 4, s(g.get("Tipo expediente", "")))
        ws.cell(r, 5, s(g.get("Fecha inicio", "")))
        ws.cell(r, 6, s(g.get("Vencimiento cuatrienal", "")))
        ws.cell(r, 7, s(g.get("Ley", "")))
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 40

    # Hoja General
    ws_g = wb.create_sheet("General")
    enc(ws_g, ["Expediente", "Campo", "Valor"])
    r = 2
    for p in proyectos:
        for campo, val in p.get("general", {}).items():
            ws_g.cell(r, 1, s(p.get("numero_expediente", "")))
            ws_g.cell(r, 2, s(campo))
            ws_g.cell(r, 3, s(val))
            r += 1
    ws_g.column_dimensions["B"].width = 35
    ws_g.column_dimensions["C"].width = 55

    # Hoja Tramitación
    ws_t = wb.create_sheet("Tramitación")
    enc(ws_t, ["Expediente", "Órgano", "Descripción", "Fecha Inicio", "Fecha Término"])
    r = 2
    for p in proyectos:
        for t in p.get("tramitacion", []):
            ws_t.cell(r, 1, s(p.get("numero_expediente", "")))
            ws_t.cell(r, 2, s(t.get("Órgano", "")))
            ws_t.cell(r, 3, s(t.get("Descripción", "")))
            ws_t.cell(r, 4, s(t.get("Fecha Inicio", "")))
            ws_t.cell(r, 5, s(t.get("Fecha Término", "")))
            r += 1
    ws_t.column_dimensions["C"].width = 50

    # Hoja Proponentes
    ws_p = wb.create_sheet("Proponentes")
    enc(ws_p, ["Expediente", "Firma", "Nombre", "Administración"])
    r = 2
    for p in proyectos:
        for prop in p.get("proponentes", []):
            ws_p.cell(r, 1, s(p.get("numero_expediente", "")))
            ws_p.cell(r, 2, s(prop.get("Firma", "")))
            ws_p.cell(r, 3, s(prop.get("Nombre", "")))
            ws_p.cell(r, 4, s(prop.get("Administración", "")))
            r += 1
    ws_p.column_dimensions["C"].width = 40

    wb.save(nombre)
    log(f"Excel guardado: {nombre}")


# ──────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────

async def main():
    inicio = datetime.now()

    log("=" * 55)
    log("FASE 1 — Extractor SIL · Asamblea Legislativa CR")
    log("=" * 55)
    log(f"Inicio:       {inicio:%Y-%m-%d %H:%M:%S}")
    log(f"Páginas:      1 a {PAGINAS_A_EXTRAER}")
    log(f"Entorno CI:   {IS_CI}")
    log("=" * 55)

    proyectos = []

    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=IS_CI,
                args=[
                    "--no-sandbox",
                    "--disable-dev-shm-usage",
                    "--disable-blink-features=AutomationControlled",
                ]
            )
            context = await browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1440, "height": 900},
            )
            page = await context.new_page()

            # Carga inicial
            log("Cargando portal SIL...")
            try:
                await page.goto(URL_BASE, wait_until="networkidle", timeout=60_000)
            except Exception as e:
                log(f"Advertencia en carga inicial: {e}")
                await page.goto(URL_BASE, wait_until="domcontentloaded", timeout=30_000)

            await page.wait_for_timeout(8_000)

            # Navegar al módulo de expedientes
            if not await navegar_a_expedientes(page):
                log("ERROR: No se pudo navegar al módulo. Saliendo limpiamente.")
                await browser.close()
                sys.exit(0)  # Salida limpia — no falla el workflow

            await page.wait_for_timeout(3_000)

            # Localizar la grilla
            frame = await encontrar_frame_con_grilla(page)
            if not frame:
                log("ERROR: Grilla no encontrada. Saliendo limpiamente.")
                await browser.close()
                sys.exit(0)

            # Configurar 10 registros por página
            await cambiar_registros_por_pagina(frame, page, REGISTROS_POR_PAG)
            await page.wait_for_timeout(4_000)

            # Asegurarse de empezar en página 1
            log("Posicionando en página 1...")
            await ir_a_pagina_1(frame, page)
            await page.wait_for_timeout(2_000)

            # ── Loop de páginas 1 a PAGINAS_A_EXTRAER ─────────────────
            for num_pagina in range(1, PAGINAS_A_EXTRAER + 1):
                exitoso = await procesar_pagina(page, frame, num_pagina, proyectos)

                if not exitoso:
                    log(f"Error en página {num_pagina}. Deteniendo Fase 1.")
                    break

                if num_pagina < PAGINAS_A_EXTRAER:
                    log(f"Avanzando a página {num_pagina + 1}...")
                    if not await ir_siguiente_pagina(frame, page, num_pagina):
                        log("No se pudo avanzar de página. Deteniendo Fase 1.")
                        break

            await browser.close()

    except Exception as e:
        log(f"Error inesperado: {e}")
        log("Continuando con lo que se extrajo hasta ahora...")

    # ── Sync y exportación ─────────────────────────────────────────────
    if not proyectos:
        log("No se extrajeron proyectos. El portal puede estar caído.")
        log("Fase 1 finalizada sin datos — sin errores en el workflow.")
        sys.exit(0)

    log("─" * 55)
    log("SINCRONIZACIÓN CON BASE DE DATOS")
    log("─" * 55)
    crear_tablas()
    stats = sync_proyectos(proyectos)

    ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_file  = f"fase1_{ts}.json"
    excel_file = f"fase1_{ts}.xlsx"

    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(proyectos, f, ensure_ascii=False, indent=2)
    log(f"JSON guardado: {json_file}")

    exportar_excel(proyectos, excel_file)

    duracion = datetime.now() - inicio

    log("=" * 55)
    log("RESUMEN FASE 1")
    log("=" * 55)
    log(f"Proyectos extraídos:  {len(proyectos)}")
    log(f"Páginas procesadas:   hasta {PAGINAS_A_EXTRAER}")
    log(f"DB sincronizados:     {stats.get('actualizados', 0)}")
    log(f"DB errores:           {stats.get('errores', 0)}")
    log(f"Duración total:       {str(duracion).split('.')[0]}")
    log("=" * 55)


if __name__ == "__main__":
    asyncio.run(main())